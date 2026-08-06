# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Finance API routes.

Endpoints:
    GET    /                    - List invoices with filters
    POST   /                    - Create invoice (auth required)
    GET    /invoices/export      - Export invoices as Excel
    GET    /invoices/{id}/br-pdf - Brazilian-styled invoice PDF (RPS layout)
    GET    /payments             - List payments
    POST   /payments             - Create payment (auth required)
    GET    /budgets              - List budgets
    POST   /budgets              - Create budget (auth required)
    PATCH  /budgets/{id}         - Update budget (auth required)
    POST   /budgets/import/file  - Import budgets from Excel/CSV (auth required)
    GET    /budgets/export       - Export budgets as Excel
    GET    /evm                  - List EVM snapshots
    POST   /evm/snapshot         - Create EVM snapshot (auth required)
    GET    /{id}                - Get single invoice
    PATCH  /{id}                - Update invoice (auth required)
    POST   /{id}/approve        - Approve invoice (auth required)
    POST   /{id}/pay            - Mark invoice as paid (auth required)

NOTE: Fixed-path routes (/payments, /budgets, /evm, /invoices/export) are
registered BEFORE the parametric /{invoice_id} route so that FastAPI does not
try to parse those path segments as UUIDs.
"""

import csv
import io
import logging
import uuid
from collections.abc import Iterable
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.file_signature import (
    SIGNATURE_BYTES_REQUIRED,
    FileSignatureMismatch,
)
from app.core.file_signature import (
    require as require_signature,
)
from app.core.rate_limiter import approval_limiter
from app.core.upload_guards import reject_if_xlsx_bomb
from app.dependencies import (
    CurrentUserId,
    RequirePermission,
    SessionDep,
    accessible_project_ids,
    verify_project_access,
)
from app.modules.contacts.models import Contact
from app.modules.finance.connector_schemas import (
    ConnectorConfigCreate,
    ConnectorConfigListResponse,
    ConnectorConfigResponse,
    ConnectorConfigUpdate,
    ConnectorTypeInfo,
    SyncLogListResponse,
    SyncLogResponse,
    SyncTriggerRequest,
    ValidateResponse,
)
from app.modules.finance.connector_service import ConnectorService
from app.modules.finance.connectors.registry import connector_registry
from app.modules.finance.models import EVMSnapshot, Invoice, Payment, ProjectBudget
from app.modules.finance.retention_ledger import RetentionRollup
from app.modules.finance.schemas import (
    BalanceSheetResponse,
    BudgetCreate,
    BudgetListResponse,
    BudgetResponse,
    BudgetUpdate,
    CashFlowResponse,
    ClaimInvoiceRequest,
    EVMListResponse,
    EVMSnapshotCreate,
    EVMSnapshotResponse,
    IncomeStatementResponse,
    InvoiceCreate,
    InvoiceListResponse,
    InvoiceResponse,
    InvoiceUpdate,
    JournalEntryCreate,
    JournalEntryResponse,
    LedgerAccountCreate,
    LedgerAccountListResponse,
    LedgerAccountResponse,
    LedgerAccountUpdate,
    LedgerEntryResponse,
    LedgerTransactionResponse,
    PaymentCreate,
    PaymentListResponse,
    PaymentResponse,
    RecordClaimPaymentRequest,
    RetentionLedgerResponse,
    RetentionRollupResponse,
    StatementLineResponse,
    TrialBalanceResponse,
    TrialBalanceRow,
)
from app.modules.finance.service import FinanceService

router = APIRouter(tags=["finance"])
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> FinanceService:
    return FinanceService(session)


# ── Counterparty enrichment ─────────────────────────────────────────────────


def _contact_display_name(c: Contact) -> str:
    """Return the human-readable contact label (company > "first last" > email)."""
    if c.company_name:
        return c.company_name
    full = f"{c.first_name or ''} {c.last_name or ''}".strip()
    return full or c.email or ""


async def _fetch_counterparty_names(session: AsyncSession, contact_ids: Iterable[str | None]) -> dict[str, str]:
    """Resolve Invoice.contact_id → display name in one round trip."""
    ids = {cid for cid in contact_ids if cid}
    if not ids:
        return {}
    rows = (await session.execute(select(Contact).where(Contact.id.in_(ids)))).scalars().all()
    return {str(c.id): _contact_display_name(c) for c in rows}


def _invoice_to_response(invoice: Invoice, names: dict[str, str]) -> InvoiceResponse:
    resp = InvoiceResponse.model_validate(invoice)
    if invoice.contact_id:
        resp.counterparty_name = names.get(invoice.contact_id)
    return resp


# ── IDOR protection helpers ─────────────────────────────────────────────────


async def _require_project_access(
    session: AsyncSession,
    project_id: uuid.UUID | None,
    user_id: str | None,
) -> None:
    """Verify the current user owns (or is admin on) the referenced project.

    Central choke-point for project-scoped finance endpoints - must be called
    before reading or writing invoices/budgets/payments/EVM snapshots that
    belong to a specific project. Mirrors the pattern used by
    ``erp_chat.tools._require_project_access`` and the shared
    :func:`app.dependencies.verify_project_access`.

    R7 hardening (2026-05-24): cross-tenant fetches now answer **404**
    rather than 403. A 403 leaks the existence of project UUIDs the
    caller is not allowed to see - the global R7 standard (and the
    shared ``verify_project_access`` helper) returns 404 on both
    "missing" and "access denied". Bringing finance in line closes the
    enumeration sidechannel for project IDs.

    A ``None`` ``project_id`` is treated as a no-op (dashboard/list
    fall-throughs that legitimately aggregate across the user's own
    projects scope it themselves in the service layer).
    """
    if project_id is None:
        return
    if user_id is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required",
        )

    try:
        from app.modules.projects.repository import ProjectRepository
        from app.modules.users.repository import UserRepository

        proj_repo = ProjectRepository(session)
        project = await proj_repo.get_by_id(project_id)
        if project is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )

        # Admin bypass
        try:
            user_repo = UserRepository(session)
            user = await user_repo.get_by_id(user_id)
            if user is not None and getattr(user, "role", "") == "admin":
                return
        except Exception:  # noqa: BLE001 - best-effort admin check
            pass

        if str(getattr(project, "owner_id", "")) != str(user_id):
            # Not the owner: allow project team members (a TeamMembership row),
            # mirroring the projects-module access guard. Otherwise 404 (R7:
            # never confirm a project UUID exists to a caller without access).
            # project_id is non-None here (guarded at the top of the function).
            from app.modules.teams.access import is_project_member

            try:
                uid = uuid.UUID(str(user_id))
            except (ValueError, TypeError):
                uid = None
            if uid is None or not await is_project_member(session, project_id, uid):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Project not found",
                )
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.warning("Finance project access check failed for %s: %s", project_id, exc)
        # Generic auth failure stays as 404 too - anything else would
        # again let the caller distinguish "exists but I lack access"
        # from "doesn't exist".
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )


async def _require_gl_consolidated_scope(
    session: AsyncSession,
    project_id: uuid.UUID | None,
    user_id: str | None,
) -> None:
    """Guard the consolidated (cross-project) GL statement endpoints.

    The GAAP statement derivations (trial balance, income statement,
    balance sheet, cash flow) accept ``project_id=None`` to aggregate the
    whole workspace ledger. Because ledger entries are not owner-scoped,
    that consolidated view spans every project on the instance, so a plain
    ``finance.gl.read`` holder must NOT be able to read it across other
    people's projects. When a project is supplied we fall back to the
    normal per-project access check; when it is omitted we require the
    caller to be an admin (the only role legitimately entitled to the
    instance-wide books). Non-admins must pass an explicit ``project_id``.
    """
    if project_id is not None:
        await _require_project_access(session, project_id, user_id)
        return
    if user_id is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required",
        )
    try:
        from app.modules.users.repository import UserRepository

        user = await UserRepository(session).get_by_id(user_id)
    except Exception:  # noqa: BLE001 - best-effort role lookup
        user = None
    if user is not None and getattr(user, "role", "") == "admin":
        return
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="A project_id is required for consolidated financial statements.",
    )


async def _require_invoice_access(
    session: AsyncSession,
    invoice_id: uuid.UUID,
    user_id: str | None,
) -> Invoice:
    """Load an invoice and verify the caller has access to its parent project."""
    invoice = await session.get(Invoice, invoice_id)
    if invoice is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Invoice {invoice_id} not found",
        )
    await _require_project_access(session, invoice.project_id, user_id)
    return invoice


async def _require_budget_access(
    session: AsyncSession,
    budget_id: uuid.UUID,
    user_id: str | None,
) -> ProjectBudget:
    """Load a budget and verify the caller has access to its parent project."""
    budget = await session.get(ProjectBudget, budget_id)
    if budget is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Budget {budget_id} not found",
        )
    await _require_project_access(session, budget.project_id, user_id)
    return budget


async def _require_payment_access(
    session: AsyncSession,
    payment_id: uuid.UUID,
    user_id: str | None,
) -> Payment:
    """Load a payment and verify caller has access via its parent invoice."""
    payment = await session.get(Payment, payment_id)
    if payment is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Payment {payment_id} not found",
        )
    await _require_invoice_access(session, payment.invoice_id, user_id)
    return payment


async def _require_evm_access(
    session: AsyncSession,
    snapshot_id: uuid.UUID,
    user_id: str | None,
) -> EVMSnapshot:
    """Load an EVM snapshot and verify caller has access to its project."""
    snapshot = await session.get(EVMSnapshot, snapshot_id)
    if snapshot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"EVM snapshot {snapshot_id} not found",
        )
    await _require_project_access(session, snapshot.project_id, user_id)
    return snapshot


# ── Invoices (list / create) ───────────────────────────────────────────────


@router.get(
    "/",
    response_model=InvoiceListResponse,
    summary="List invoices",
    description="Retrieve a paginated list of invoices with optional filters by project, "
    "direction (payable/receivable), and status.",
)
async def list_invoices(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    direction: str | None = Query(default=None),
    status: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceListResponse:
    """List invoices with optional filters."""
    await _require_project_access(session, project_id, user_id)
    # When no project is named, scope the listing to the caller's accessible
    # projects so a VIEWER does not receive every tenant's invoices. Admins get
    # None (no filter -> full portfolio view).
    scope = None if project_id is not None else await accessible_project_ids(session, user_id)
    items, total = await service.list_invoices(
        project_id=project_id,
        project_ids=scope,
        direction=direction,
        invoice_status=status,
        offset=offset,
        limit=limit,
    )
    names = await _fetch_counterparty_names(session, (i.contact_id for i in items))
    return InvoiceListResponse(
        items=[_invoice_to_response(i, names) for i in items],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.post(
    "/",
    response_model=InvoiceResponse,
    status_code=201,
    summary="Create invoice",
    description="Create a new invoice with optional line items. Set invoice_direction "
    "to 'payable' (vendor invoices) or 'receivable' (client invoices).",
)
async def create_invoice(
    data: InvoiceCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Create a new invoice."""
    await _require_project_access(session, data.project_id, user_id)
    invoice = await service.create_invoice(data, user_id=user_id)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


# ── /invoices alias collection (BUG-API12) ─────────────────────────────────
#
# Frontend and external clients expect ``/api/v1/finance/invoices`` (without
# trailing slash) and ``/api/v1/finance/invoices/`` to behave like the root
# ``/``. Without an explicit alias FastAPI matches the ``/{invoice_id}``
# parametric route below and returns 422 on the literal string "invoices"
# (UUID parse error). Declared BEFORE ``/{invoice_id}`` so the static path
# wins.


@router.get(
    "/invoices/",
    response_model=InvoiceListResponse,
    summary="List invoices (alias of GET /)",
    description="Alias of ``GET /api/v1/finance/`` - returns the same paginated "
    "list. Provided so that clients hitting ``/api/v1/finance/invoices`` get a "
    "sensible 200 instead of a UUID-parse 422.",
)
async def list_invoices_alias(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    direction: str | None = Query(default=None),
    status: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceListResponse:
    """List invoices (alias of ``GET /``).

    Replicates the behaviour of ``list_invoices`` inline (rather than calling
    it through Python) so that FastAPI dependency wiring resolves cleanly.
    """
    await _require_project_access(session, project_id, user_id)
    # Mirror list_invoices: scope an unfiltered listing to the caller's
    # accessible projects (admins -> None -> full portfolio).
    scope = None if project_id is not None else await accessible_project_ids(session, user_id)
    items, total = await service.list_invoices(
        project_id=project_id,
        project_ids=scope,
        direction=direction,
        invoice_status=status,
        offset=offset,
        limit=limit,
    )
    names = await _fetch_counterparty_names(session, (i.contact_id for i in items))
    return InvoiceListResponse(
        items=[_invoice_to_response(i, names) for i in items],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── Export invoices as Excel ────────────────────────────────────────────────


@router.get(
    "/invoices/export/",
    summary="Export invoices as Excel",
    description="Download invoices for a project as an Excel (.xlsx) file. "
    "Optionally filter by direction (payable/receivable).",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_invoices(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    direction: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
) -> StreamingResponse:
    """Export invoices for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    await _require_project_access(session, project_id, _user_id)

    stmt = select(Invoice).where(Invoice.project_id == project_id)
    if direction:
        stmt = stmt.where(Invoice.invoice_direction == direction)
    stmt = stmt.limit(50000)

    result = await session.execute(stmt)
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Invoice #",
        "Direction",
        "Date",
        "Due Date",
        "Vendor/Client",
        "Subtotal",
        "Tax",
        "Total",
        "Status",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    from decimal import Decimal, InvalidOperation

    def _safe_decimal(raw: Any) -> Decimal:
        """Coerce DB string to Decimal preserving precision; NaN/Inf → 0.

        openpyxl accepts Decimal natively and stores as number without the
        float-precision loss ``float()`` would introduce on large currency
        values (BUG-069/070).
        """
        if raw is None or raw == "":
            return Decimal("0")
        try:
            d = Decimal(str(raw).strip())
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0")
        return d if d.is_finite() else Decimal("0")

    for row_idx, inv in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=inv.invoice_number)
        ws.cell(row=row_idx, column=2, value=inv.invoice_direction)
        ws.cell(row=row_idx, column=3, value=inv.invoice_date)
        ws.cell(row=row_idx, column=4, value=inv.due_date)
        ws.cell(row=row_idx, column=5, value=inv.contact_id or "")
        ws.cell(row=row_idx, column=6, value=_safe_decimal(inv.amount_subtotal))
        ws.cell(row=row_idx, column=7, value=_safe_decimal(inv.tax_amount))
        ws.cell(row=row_idx, column=8, value=_safe_decimal(inv.amount_total))
        ws.cell(row=row_idx, column=9, value=inv.status)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices_export.xlsx"'},
    )


# ── Brazilian-styled invoice PDF (Tier-1 - pre-NF-e bridge) ─────────────────
#
# Path lives under ``/invoices/{invoice_id}/br-pdf/`` so FastAPI's static
# prefix ``/invoices/`` wins over the bare ``/{invoice_id}`` parametric
# route. See ``br_invoice_pdf.py`` for the rendering logic and the
# disclaimer text explaining why this PDF is NOT a fiscal document
# (NF-e / NFS-e SEFAZ integration is Tier-2 - see
# ``__brazil_tier2_followups.md``).


@router.get(
    "/invoices/{invoice_id}/br-pdf/",
    summary="Export invoice as Brazil-styled PDF (RPS layout)",
    description=(
        "Render the invoice as a one-page PDF in the Brazilian RPS "
        "(Recibo Provisório de Serviços) layout, with CNPJ / IE / Razão "
        "Social / código de serviço / retenções fields. NOT a fiscal "
        "document - for full NF-e / NFS-e SEFAZ output see Tier-2 roadmap."
    ),
    response_description="application/pdf stream",
)
async def export_invoice_br_pdf(
    invoice_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> StreamingResponse:
    """Render a Brazilian-styled invoice PDF and stream it back."""
    from app.modules.finance.br_invoice_pdf import render_br_invoice_pdf

    invoice = await _require_invoice_access(session, invoice_id, user_id)
    fresh = await service.get_invoice(invoice_id)

    # Project context (best-effort - never block the PDF on project lookup)
    project_dict: dict[str, Any] = {}
    try:
        from app.modules.projects.repository import ProjectRepository

        proj = await ProjectRepository(session).get_by_id(fresh.project_id)
        if proj is not None:
            project_dict = {
                "name": getattr(proj, "name", "") or "",
                "code": getattr(proj, "code", "") or "",
            }
    except Exception:  # noqa: BLE001 - header is decorative
        logger.debug("BR invoice PDF: project lookup failed", exc_info=True)

    invoice_dict: dict[str, Any] = {
        "invoice_number": fresh.invoice_number,
        "invoice_direction": fresh.invoice_direction,
        "invoice_date": fresh.invoice_date,
        "due_date": fresh.due_date,
        "amount_subtotal": fresh.amount_subtotal,
        "tax_amount": fresh.tax_amount,
        "retention_amount": fresh.retention_amount,
        "amount_total": fresh.amount_total,
        "notes": fresh.notes,
        "metadata": dict(fresh.metadata_ or {}),
    }
    line_items: list[dict[str, Any]] = [
        {
            "description": li.description,
            "unit": li.unit,
            "quantity": li.quantity,
            "unit_rate": li.unit_rate,
            "amount": li.amount,
        }
        for li in (fresh.line_items or [])
    ]

    pdf_bytes = render_br_invoice_pdf(
        invoice=invoice_dict,
        line_items=line_items,
        project=project_dict or None,
    )

    # Sanitise invoice_number before embedding in a quoted Content-Disposition
    # header.  invoice_number is a user-controlled DB value - it can contain
    # characters that would break the RFC 6266 quoted-string or inject
    # additional headers (CRLF injection).  Strip every character that is not
    # ASCII printable, remove double-quotes (which terminate the quoted-string
    # token) and forward-slashes (already done historically), and cap length.
    _raw_num = invoice.invoice_number or "invoice"
    _safe_num = (
        (
            _raw_num.encode("ascii", errors="replace")  # non-ASCII → b'?'
            .decode("ascii")
            .replace("\r", "")
            .replace("\n", "")
            .replace('"', "'")
            .replace("/", "-")
            .strip()
        )[:80]
        or "invoice"
    )
    filename = f"RPS_{_safe_num}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/invoices/{invoice_id}/einvoice",
    summary="Export invoice as an EN 16931 e-invoice (international: CII and UBL/Peppol)",
    description=(
        "Render the invoice as an EN 16931 electronic invoice. EN 16931 is the "
        "international semantic standard, so the same invoice can be issued in "
        "any supported country flavour and either syntax: CII for ZUGFeRD 2.1, "
        "Factur-X 1.0 and XRechnung 3.0 (DACH/EU), or UBL for Peppol BIS "
        "Billing 3.0 and plain EN 16931 UBL (worldwide - EU, UK, Australia, New "
        "Zealand, Singapore and more). Choose with "
        "?format=peppol|ubl|xrechnung|zugferd|facturx|en16931. Seller and buyer "
        "master data, the Buyer reference (Leitweg-ID for XRechnung, PO for "
        "Peppol) and an explicit VAT rate are read from the invoice metadata "
        "under the 'einvoice' key. Pass ?dry_run=true to get the list of missing "
        "EN 16931 fields as JSON instead of the file, so the UI can prompt for them."
    ),
    response_description="application/xml e-invoice stream, or a JSON problem list when dry_run=true",
    response_model=None,
)
async def export_invoice_einvoice(
    invoice_id: uuid.UUID,
    session: SessionDep,
    fmt: str = Query(default="xrechnung", alias="format"),
    dry_run: bool = Query(default=False),
    embed: bool = Query(default=False),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> StreamingResponse | dict[str, Any]:
    """Stream an EN 16931 e-invoice (CII/UBL XML, or a hybrid PDF) from the invoice."""
    from app.modules.einvoice import (
        SUPPORTED_PROFILES,
        problems_for,
        render_einvoice,
        render_einvoice_pdf,
    )
    from app.modules.einvoice.cii import EInvoiceError

    profile = (fmt or "xrechnung").strip().lower()
    if profile not in SUPPORTED_PROFILES:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"unknown e-invoice format {fmt!r}; use one of {', '.join(SUPPORTED_PROFILES)}",
        )

    await _require_invoice_access(session, invoice_id, user_id)
    fresh = await service.get_invoice(invoice_id)

    # Best-effort buyer name fallback from the linked contact (never block on it).
    buyer_fallback = ""
    if fresh.contact_id:
        try:
            from app.modules.contacts.repository import ContactRepository

            contact = await ContactRepository(session).get_by_id(fresh.contact_id)
            if contact is not None:
                buyer_fallback = str(getattr(contact, "name", "") or "").strip()
        except Exception:  # noqa: BLE001 - fallback only
            logger.debug("e-invoice: contact lookup failed", exc_info=True)

    invoice_dict: dict[str, Any] = {
        "invoice_number": fresh.invoice_number,
        "invoice_direction": fresh.invoice_direction,
        "invoice_date": fresh.invoice_date,
        "due_date": fresh.due_date,
        "currency_code": fresh.currency_code,
        "amount_subtotal": fresh.amount_subtotal,
        "tax_amount": fresh.tax_amount,
        "retention_amount": fresh.retention_amount,
        "amount_total": fresh.amount_total,
        "notes": fresh.notes,
        "metadata": dict(fresh.metadata_ or {}),
    }
    line_items: list[dict[str, Any]] = [
        {
            "description": li.description,
            "unit": li.unit,
            "quantity": li.quantity,
            "unit_rate": li.unit_rate,
            "amount": li.amount,
        }
        for li in (fresh.line_items or [])
    ]

    if dry_run:
        problems = problems_for(
            invoice=invoice_dict,
            line_items=line_items,
            profile=profile,
            buyer_fallback_name=buyer_fallback,
        )
        return {"format": profile, "valid": not problems, "problems": problems}

    render = render_einvoice_pdf if embed else render_einvoice
    try:
        filename, media_type, body = render(
            invoice=invoice_dict,
            line_items=line_items,
            profile=profile,
            buyer_fallback_name=buyer_fallback,
        )
    except EInvoiceError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"invoice is not EN 16931 complete for {profile}: {exc}. "
                "Fill seller/buyer master data and the buyer reference under the "
                "invoice metadata 'einvoice' key, or call with ?dry_run=true."
            ),
        ) from exc

    # ``filename`` is already ASCII-sanitised by the service (_safe_token).
    return StreamingResponse(
        io.BytesIO(body),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Payments (MUST be before /{invoice_id}) ─────────────────────────────────


@router.get(
    "/payments/",
    response_model=PaymentListResponse,
    summary="List payments",
    description="Retrieve a paginated list of payments, optionally filtered by invoice.",
)
async def list_payments(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    invoice_id: uuid.UUID | None = Query(default=None),
    project_id: uuid.UUID | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=500),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> PaymentListResponse:
    """List payments scoped to a single invoice or project.

    Either ``invoice_id`` or ``project_id`` MUST be supplied - an unscoped
    call would return payments across every tenant. Access to the referenced
    invoice/project is verified before any rows are read.
    """
    if invoice_id is not None:
        await _require_invoice_access(session, invoice_id, user_id)
    elif project_id is not None:
        await _require_project_access(session, project_id, user_id)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either invoice_id or project_id is required to list payments.",
        )
    items, total = await service.list_payments(
        invoice_id=invoice_id,
        project_id=project_id,
        limit=limit,
        offset=offset,
    )

    # Enrich each payment with its parent invoice number (one round trip)
    # so the UI can show a readable reference rather than a raw UUID.
    invoice_ids = {p.invoice_id for p in items}
    numbers: dict[uuid.UUID, str] = {}
    if invoice_ids:
        rows = (
            await session.execute(select(Invoice.id, Invoice.invoice_number).where(Invoice.id.in_(invoice_ids)))
        ).all()
        numbers = {row[0]: row[1] for row in rows}

    responses: list[PaymentResponse] = []
    for p in items:
        resp = PaymentResponse.model_validate(p)
        resp.invoice_number = numbers.get(p.invoice_id)
        resp.status = "refunded" if p.is_refund else "completed"
        responses.append(resp)

    return PaymentListResponse(items=responses, total=total)


@router.post(
    "/payments/",
    response_model=PaymentResponse,
    status_code=201,
    summary="Create payment",
    description="Record a payment against an invoice. Updates the invoice's paid amount. "
    "MANAGER-only - recording a payment is a binding ledger entry, not a CRUD edit.",
)
async def create_payment(
    data: PaymentCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.record_payment")),
    service: FinanceService = Depends(_get_service),
) -> PaymentResponse:
    """Record a payment against an invoice.

    R7 (2026-05-24): pinned to ``finance.record_payment`` (MANAGER+).
    Recording a payment row is a financial commitment that affects the
    invoice's paid/outstanding state and downstream budget actuals - it
    cannot remain an EDITOR-level action.
    """
    await _require_invoice_access(session, data.invoice_id, user_id)
    payment = await service.create_payment(data, actor_id=str(user_id) if user_id else None)
    return PaymentResponse.model_validate(payment)


# ── THCC local payment std import (MUST be before /{invoice_id}) ─────────────


@router.get(
    "/payments/thcc/config/",
    summary="THCC payment std root path",
    description="Return the local folder scanned for B_财务付款_std Excel files.",
)
async def thcc_payment_config(
    _user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.read")),
) -> dict:
    from app.modules.finance.thcc_payment_import import get_payments_root

    root = get_payments_root()
    return {
        "root": str(root),
        "exists": root.is_dir(),
        "default_hint": (
            "~/Desktop/邯郸中材/01成本统计/12-财务数据💰/A_财务付款数据/B_财务付款_std"
        ),
    }


@router.put(
    "/payments/thcc/config/",
    summary="Set THCC payment std root path",
)
async def thcc_payment_set_config(
    body: dict[str, Any],
    _user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.record_payment")),
) -> dict:
    from app.modules.finance.thcc_payment_import import get_payments_root, set_payments_root

    raw = (body.get("root") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="root path is required")
    path = set_payments_root(raw)
    return {"root": str(path), "exists": path.is_dir()}


@router.get(
    "/payments/thcc/scan/",
    summary="Scan local THCC payment std folder",
    description="List Excel files under the configured B_财务付款_std path and "
    "match them to OCE projects by project_code / name.",
)
async def thcc_payment_scan(
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.read")),
) -> dict:
    from app.modules.finance.thcc_payment_import import (
        enrich_scan_with_projects,
        get_payments_root,
        scan_payment_files,
    )

    root = get_payments_root()
    if not root.is_dir():
        return {
            "root": str(root),
            "exists": False,
            "files": [],
            "error": f"Directory not found: {root}",
        }
    files = scan_payment_files(root)
    enriched = await enrich_scan_with_projects(session, files)
    return {
        "root": str(root),
        "exists": True,
        "file_count": len(enriched),
        "matched": sum(1 for f in enriched if f.get("project_matched")),
        "files": enriched,
    }


@router.post(
    "/payments/thcc/import/",
    summary="Bulk-import THCC payment std ledgers",
    description="Import payment rows from local B_财务付款_std Excel files. "
    "Creates one paid payable invoice + payment per 实付金额 row. "
    "Re-import is idempotent (stable thccpay-* keys). "
    "Optional filters: project_id (current project), project_codes, filenames, dry_run.",
)
async def thcc_payment_import(
    body: dict[str, Any],
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.record_payment")),
) -> dict:
    from app.modules.finance.thcc_payment_import import import_from_local_root

    project_id_raw = body.get("project_id")
    project_id: uuid.UUID | None = None
    if project_id_raw:
        try:
            project_id = uuid.UUID(str(project_id_raw))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Invalid project_id") from exc
        await _require_project_access(session, project_id, user_id)

    codes = body.get("project_codes") or None
    if codes is not None and not isinstance(codes, list):
        raise HTTPException(status_code=400, detail="project_codes must be a list")
    filenames = body.get("filenames") or None
    if filenames is not None and not isinstance(filenames, list):
        raise HTTPException(status_code=400, detail="filenames must be a list")
    dry_run = bool(body.get("dry_run") or False)

    result = await import_from_local_root(
        session,
        project_id=project_id,
        project_codes=[str(c) for c in codes] if codes else None,
        filenames=[str(f) for f in filenames] if filenames else None,
        dry_run=dry_run,
        actor_id=str(user_id) if user_id else None,
    )
    return result


@router.get(
    "/budgets/thcc-resp-cost/config/",
    summary="THCC responsibility-cost std root path",
)
async def thcc_resp_cost_config(
    _user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.read")),
) -> dict:
    from app.modules.finance.thcc_resp_cost_import import get_resp_cost_root

    root = get_resp_cost_root()
    return {
        "root": str(root),
        "exists": root.is_dir(),
        "default_hint": "~/Desktop/邯郸中材/01成本统计/3.1_责任成本/2_责任成本_std",
    }


@router.put(
    "/budgets/thcc-resp-cost/config/",
    summary="Set THCC responsibility-cost std root path",
)
async def thcc_resp_cost_set_config(
    body: dict[str, Any],
    _user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.update")),
) -> dict:
    from app.modules.finance.thcc_resp_cost_import import set_resp_cost_root

    raw = (body.get("root") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="root path is required")
    path = set_resp_cost_root(raw)
    return {"root": str(path), "exists": path.is_dir()}


@router.get(
    "/budgets/thcc-resp-cost/scan/",
    summary="Scan local THCC responsibility-cost std folder",
    description="List 责任成本_std Excel files and match them to OCE projects by project_code.",
)
async def thcc_resp_cost_scan(
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.read")),
) -> dict:
    from app.modules.finance.thcc_resp_cost_import import (
        enrich_scan,
        get_resp_cost_root,
        scan_resp_cost_files,
    )

    root = get_resp_cost_root()
    if not root.is_dir():
        return {
            "root": str(root),
            "exists": False,
            "files": [],
            "error": f"Directory not found: {root}",
        }
    files = scan_resp_cost_files(root)
    enriched = await enrich_scan(session, files)
    return {
        "root": str(root),
        "exists": True,
        "file_count": len(enriched),
        "matched": sum(1 for f in enriched if f.get("project_matched")),
        "files": enriched,
    }


@router.post(
    "/budgets/thcc-resp-cost/import/",
    summary="Bulk-import THCC responsibility cost as budget lines",
    description=(
        "Import leaf lines from 2_责任成本_std workbooks into oe_finance_budget. "
        "Each A.1/B.2… line becomes a budget row (wbs_id=code, category=name). "
        "CNY amounts convert to THB at fx_cny_to_thb (default 4.9). "
        "Re-import upserts; replace=true drops stale tagged lines. "
        "Also refreshes cost-board snapshot resp_cost when available."
    ),
)
async def thcc_resp_cost_import(
    body: dict[str, Any],
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.create")),
) -> dict:
    from decimal import Decimal, InvalidOperation

    from app.modules.finance.thcc_resp_cost_import import import_from_local_root

    project_id_raw = body.get("project_id")
    project_id: uuid.UUID | None = None
    if project_id_raw:
        try:
            project_id = uuid.UUID(str(project_id_raw))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Invalid project_id") from exc
        await _require_project_access(session, project_id, user_id)

    codes = body.get("project_codes") or None
    if codes is not None and not isinstance(codes, list):
        raise HTTPException(status_code=400, detail="project_codes must be a list")
    filenames = body.get("filenames") or None
    if filenames is not None and not isinstance(filenames, list):
        raise HTTPException(status_code=400, detail="filenames must be a list")
    dry_run = bool(body.get("dry_run") or False)
    replace = body.get("replace", True)
    if not isinstance(replace, bool):
        replace = bool(replace)

    fx = None
    if body.get("fx_cny_to_thb") is not None:
        try:
            fx = Decimal(str(body["fx_cny_to_thb"]))
        except (InvalidOperation, ValueError, TypeError) as exc:
            raise HTTPException(status_code=400, detail="Invalid fx_cny_to_thb") from exc

    return await import_from_local_root(
        session,
        project_id=project_id,
        project_codes=[str(c) for c in codes] if codes else None,
        filenames=[str(f) for f in filenames] if filenames else None,
        dry_run=dry_run,
        replace=replace,
        fx_cny_to_thb=fx,
        actor_id=str(user_id) if user_id else None,
        sync_cost_board=bool(body.get("sync_cost_board", True)),
    )


@router.post(
    "/budgets/thcc-resp-cost/import-upload/",
    summary="Import one THCC responsibility-cost Excel upload as budgets",
)
async def thcc_resp_cost_import_upload(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    dry_run: bool = Query(default=False),
    replace: bool = Query(default=True),
    file: UploadFile = File(...),
    _perm: None = Depends(RequirePermission("finance.create")),
) -> dict:
    from app.modules.finance.thcc_resp_cost_import import import_resp_cost_file

    await _require_project_access(session, project_id, user_id)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")
    reject_if_xlsx_bomb(content)

    import tempfile
    from pathlib import Path

    suffix = Path(file.filename or "resp.xlsx").suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=True) as tmp:
        tmp.write(content)
        tmp.flush()
        try:
            return await import_resp_cost_file(
                session,
                path=tmp.name,
                project_id=project_id,
                dry_run=dry_run,
                replace=replace,
                actor_id=str(user_id) if user_id else None,
            )
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Import failed: {exc}") from exc


@router.post(
    "/payments/thcc/import-upload/",
    summary="Import a single THCC payment std Excel upload",
    description="Upload one B_财务付款_std workbook for the given project_id.",
)
async def thcc_payment_import_upload(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    dry_run: bool = Query(default=False),
    file: UploadFile = File(...),
    _perm: None = Depends(RequirePermission("finance.record_payment")),
) -> dict:
    from app.modules.finance.thcc_payment_import import (
        import_payment_rows,
        parse_payment_workbook,
    )
    from app.modules.projects.models import Project

    await _require_project_access(session, project_id, user_id)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")
    reject_if_xlsx_bomb(content)

    import tempfile
    from pathlib import Path

    suffix = Path(file.filename or "pay.xlsx").suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=True) as tmp:
        tmp.write(content)
        tmp.flush()
        path = Path(tmp.name)
        try:
            rows = parse_payment_workbook(path)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {exc}") from exc

    # rewrite source_file for metadata
    for r in rows:
        r.source_file = file.filename or path.name

    proj = await session.get(Project, project_id)
    if proj is None:
        raise HTTPException(status_code=404, detail="Project not found")
    currency = (proj.currency or "").strip() or "THB"
    stats = await import_payment_rows(
        session,
        rows=rows,
        project_id=project_id,
        currency_code=currency,
        dry_run=dry_run,
        actor_id=str(user_id) if user_id else None,
    )
    return {
        "ok": True,
        "filename": file.filename,
        "project_id": str(project_id),
        **stats,
    }


# ── Gap E: certified claim → receivable invoice (MUST be before /{id}) ───────


@router.post(
    "/invoices/from-claim/",
    response_model=InvoiceResponse,
    summary="Create receivable invoice from a certified claim",
    description="Auto-create (or return the existing) accounts-receivable invoice for a "
    "certified progress claim. Idempotent on claim_id: a second call returns the same "
    "invoice with 200 rather than writing a duplicate. The claim must be in 'certified' "
    "status (400 otherwise). MANAGER-only - it books revenue against the client.",
)
async def create_invoice_from_claim(
    data: ClaimInvoiceRequest,
    user_id: CurrentUserId,
    session: SessionDep,
    response: Response,
    _perm: None = Depends(RequirePermission("finance.invoice_from_claim")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Create or return the receivable invoice for a certified claim.

    Returns 201 when a new invoice is created, 200 when an existing one is
    returned (idempotent retry). The 200/201 distinction is set on the
    response after the service resolves whether the invoice already existed.
    """
    pre_existing = await service.get_receivable_for_claim(data.claim_id)
    invoice = await service.create_receivable_from_claim(
        data.claim_id,
        actor_id=str(user_id) if user_id else None,
    )
    # Authorise against the resolved project (the claim's contract project).
    await _require_project_access(session, invoice.project_id, user_id)
    response.status_code = status.HTTP_200_OK if pre_existing is not None else status.HTTP_201_CREATED
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.get(
    "/claims/{claim_id}/receivable-invoice/",
    response_model=InvoiceResponse,
    summary="Get the receivable invoice raised from a claim",
    description="Convenience lookup: return the accounts-receivable invoice that was "
    "auto-created from the given certified progress claim. 404 when none exists yet.",
)
async def get_receivable_for_claim(
    claim_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Look up the receivable invoice for a claim."""
    invoice = await service.get_receivable_for_claim(claim_id)
    if invoice is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No receivable invoice exists for this claim",
        )
    await _require_project_access(session, invoice.project_id, user_id)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.post(
    "/invoices/{invoice_id}/record-payment/",
    response_model=PaymentResponse,
    status_code=201,
    summary="Record a payment with retainage withholding",
    description="Record a payment against an invoice, holding back retainage. When "
    "withholding_amount is omitted it is derived from the invoice retention_amount; when "
    "amount is omitted the invoice net (total - retention) is paid. Idempotent on "
    "idempotency_key. The cash leg (not the withheld retainage) is posted to the cost "
    "spine. MANAGER-only - a payment is a binding ledger entry.",
)
async def record_payment_with_withholding(
    invoice_id: uuid.UUID,
    data: RecordClaimPaymentRequest,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.record_payment")),
    service: FinanceService = Depends(_get_service),
) -> PaymentResponse:
    """Record a withholding payment against an invoice."""
    await _require_invoice_access(session, invoice_id, user_id)
    payment = await service.record_payment_with_withholding(
        invoice_id,
        data,
        actor_id=str(user_id) if user_id else None,
    )
    return PaymentResponse.model_validate(payment)


# -- Retention / withholding ledger (MUST be before /{invoice_id}) ------------


def _retention_money(value: Decimal) -> str:
    """Serialise an always-present retention Decimal to a canonical string."""
    return format(value, "f")


def _retention_pct(value: Decimal | None) -> str | None:
    """Serialise a guarded ratio Decimal to a string, or None when absent."""
    return None if value is None else format(value, "f")


def _retention_rollup_to_response(
    rollup: RetentionRollup,
    names: dict[str, str],
) -> RetentionRollupResponse:
    """Map a pure RetentionRollup to its API response, adding the contact name."""
    return RetentionRollupResponse(
        currency_code=rollup.currency_code,
        direction=rollup.direction,
        contact_id=rollup.contact_id,
        counterparty_name=names.get(rollup.contact_id) if rollup.contact_id else None,
        scheduled=_retention_money(rollup.scheduled),
        held_to_date=_retention_money(rollup.held_to_date),
        released_to_date=_retention_money(rollup.released_to_date),
        outstanding=_retention_money(rollup.outstanding),
        payment_count=rollup.payment_count,
        released_pct=_retention_pct(rollup.released_pct),
        outstanding_pct=_retention_pct(rollup.outstanding_pct),
        held_vs_scheduled_pct=_retention_pct(rollup.held_vs_scheduled_pct),
        earliest_release_date=rollup.earliest_release_date,
        latest_release_date=rollup.latest_release_date,
    )


@router.get(
    "/retention-ledger/",
    response_model=RetentionLedgerResponse,
    summary="Retention / withholding ledger for a project",
    description=(
        "Unified retainage rollup for a project, computed from stored invoice "
        "retention and payment withholding. Returns per-counterparty lines and "
        "per-(currency, direction) totals: retention scheduled, held to date, "
        "released to date and still outstanding. 'Released' means the "
        "contractual release date has been reached as of ?as_of (default: "
        "today); the module records no separate cash return, so this is "
        "release-due, not cash returned. Nothing is blended across currencies "
        "or payable / receivable."
    ),
)
async def get_retention_ledger(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    as_of: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> RetentionLedgerResponse:
    """Return the project retention / withholding ledger rollup."""
    await verify_project_access(project_id, user_id, session)

    ledger = await service.get_retention_ledger(project_id, as_of=as_of)

    # Resolve counterparty display names for the per-contact lines in one round
    # trip (totals carry no contact_id, so they need no name).
    names = await _fetch_counterparty_names(session, [g.contact_id for g in ledger.groups])

    return RetentionLedgerResponse(
        project_id=project_id,
        as_of=ledger.as_of,
        groups=[_retention_rollup_to_response(g, names) for g in ledger.groups],
        totals=[_retention_rollup_to_response(t, names) for t in ledger.totals],
    )


# ── Budgets (MUST be before /{invoice_id}) ──────────────────────────────────


@router.get(
    "/budgets/",
    response_model=BudgetListResponse,
    summary="List budgets",
    description="Retrieve project budget lines with optional filters by project and cost category.",
)
async def list_budgets(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    category: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> BudgetListResponse:
    """List project budgets."""
    await _require_project_access(session, project_id, user_id)
    # Scope an unfiltered listing to the caller's accessible projects so a
    # VIEWER does not receive every tenant's budgets (admins -> None -> all).
    scope = None if project_id is not None else await accessible_project_ids(session, user_id)
    items, total = await service.list_budgets(project_id=project_id, project_ids=scope, category=category)
    return BudgetListResponse(
        items=[BudgetResponse.model_validate(b) for b in items],
        total=total,
    )


@router.post(
    "/budgets/",
    response_model=BudgetResponse,
    status_code=201,
    summary="Create budget line",
    description="Create a project budget line for a specific WBS element and cost category.",
)
async def create_budget(
    data: BudgetCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Create a project budget line."""
    await _require_project_access(session, data.project_id, user_id)
    budget = await service.create_budget(data)
    return BudgetResponse.model_validate(budget)


# ── Budget import (CSV / Excel) ─────────────────────────────────────────────

_BUDGET_COLUMN_ALIASES: dict[str, list[str]] = {
    "wbs_id": [
        "wbs_id",
        "wbs code",
        "wbs",
        "code",
        "wbs_code",
    ],
    "category": [
        "category",
        "cost category",
        "kategorie",
        "type",
    ],
    "original_budget": [
        "original_budget",
        "original budget",
        "original",
        "budget",
        "amount",
    ],
    "notes": [
        "notes",
        "note",
        "remarks",
        "bemerkung",
    ],
}

_ALLOWED_BUDGET_CATEGORIES = {
    "labor",
    "material",
    "equipment",
    "subcontractor",
    "overhead",
    "contingency",
    "other",
}


def _match_budget_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _BUDGET_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _safe_decimal_str(value: Any, default: str = "0") -> str:
    """Parse a value to a decimal string, returning default on failure."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" -> "1234.56"
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        float(text)  # validate
        return text
    except (ValueError, TypeError):
        return default


def _parse_budget_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for budget import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_budget_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_budget_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for budget import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_budget_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


@router.post(
    "/budgets/import/file/",
    summary="Import budgets from file",
    description="Upload an Excel (.xlsx) or CSV (.csv) file to bulk-import budget lines. "
    "Column headers are auto-detected using flexible aliases (EN/DE). "
    "Returns a summary with imported, skipped, and error counts per row.",
)
async def import_budgets_file(
    _user_id: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> dict[str, Any]:
    """Import project budgets from an Excel or CSV file upload.

    Expected columns:
    - **WBS Code** -- work breakdown structure code
    - **Category** -- budget category (labor, material, equipment, etc.)
    - **Original Budget** -- original budget amount
    - **Notes** -- optional notes

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    await _require_project_access(session, project_id, _user_id)

    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # R7 (2026-05-24): magic-byte gate. The filename-extension check above
    # is necessary but trivially bypassable - an attacker who renames a
    # PE/ELF/script to ``payload.xlsx`` would pass the extension test and
    # land in the parser. ``require_signature`` rejects anything that
    # isn't a ZIP-container (xlsx/xls OLE) or plain-text/CSV (which has
    # no signature and surfaces as ``None``).
    #
    # CSV genuinely has no magic bytes so it returns ``None`` from
    # ``detect`` - only require the signature check for the spreadsheet
    # branches; plain CSV falls through to the parser as-is.
    fname_low = filename
    if fname_low.endswith((".xlsx", ".xls")):
        head = content[:SIGNATURE_BYTES_REQUIRED]
        try:
            require_signature(
                head,
                # xlsx → ZIP container; legacy .xls → OLE compound document.
                frozenset({"zip", "ole"}),
                filename=file.filename,
            )
        except FileSignatureMismatch as exc:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=str(exc),
            )

    # Zip-bomb guard: reject .xlsx whose uncompressed sheets exceed 50 MB.
    reject_if_xlsx_bomb(content)

    # Parse rows based on file type
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_budget_rows_from_excel(content)
        else:
            rows = _parse_budget_rows_from_csv(content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing budget import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # Convert rows to BudgetCreate objects and import
    imported_count = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            wbs_id = str(row.get("wbs_id", "")).strip() or None

            # Parse category
            category = str(row.get("category", "")).strip().lower() or None
            if category and category not in _ALLOWED_BUDGET_CATEGORIES:
                errors.append(
                    {
                        "row": row_idx,
                        "error": (
                            f"Invalid category: '{category}'. Allowed: {', '.join(sorted(_ALLOWED_BUDGET_CATEGORIES))}"
                        ),
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            # Parse amount
            original_budget = _safe_decimal_str(row.get("original_budget"))

            # Validate amount is a valid number
            try:
                float(original_budget)
            except (ValueError, TypeError):
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Invalid budget amount: {row.get('original_budget')}",
                        "data": {k: str(v)[:100] for k, v in row.items()},
                    }
                )
                continue

            # Skip rows with no data
            if not wbs_id and not category and original_budget == "0":
                skipped += 1
                continue

            data = BudgetCreate(
                project_id=project_id,
                wbs_id=wbs_id,
                category=category,
                original_budget=original_budget,
                revised_budget=original_budget,  # default revised = original
            )
            await service.create_budget(data)
            imported_count += 1

        except Exception as exc:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(exc),
                    "data": {k: str(v)[:100] for k, v in row.items()},
                }
            )
            logger.warning("Budget import error at row %d: %s", row_idx, exc)

    logger.info(
        "Budget file import complete: imported=%d, skipped=%d, errors=%d",
        imported_count,
        skipped,
        len(errors),
    )

    return {
        "imported": imported_count,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


# ── Export budgets as Excel ──────────────────────────────────────────────────


@router.get(
    "/budgets/export/",
    summary="Export budgets as Excel",
    description="Download budgets for a project as an Excel (.xlsx) file with "
    "original, revised, committed, actual, forecast, and variance columns.",
    response_description="Excel file stream (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)",
)
async def export_budgets(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(...),
    _perm: None = Depends(RequirePermission("finance.read")),
) -> StreamingResponse:
    """Export budgets for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    await _require_project_access(session, project_id, _user_id)

    result = await session.execute(select(ProjectBudget).where(ProjectBudget.project_id == project_id).limit(50000))
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Budgets"

    headers = [
        "WBS",
        "Category",
        "Original",
        "Revised",
        "Committed",
        "Actual",
        "Forecast",
        "Variance",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, b in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=b.wbs_id or "")
        ws.cell(row=row_idx, column=2, value=b.category or "")
        # BUG-069: use Decimal (not float) so large construction-budget values
        # (e.g. 123456789.99) don't suffer IEEE-754 rounding when Excel reads
        # them back - openpyxl stores Decimal natively as a NUMERIC cell.
        from decimal import Decimal as _Dec
        from decimal import InvalidOperation as _IOp

        def _bd(raw: Any) -> _Dec:
            if raw is None or raw == "":
                return _Dec("0")
            try:
                d = _Dec(str(raw).strip())
            except (_IOp, ValueError, TypeError):
                return _Dec("0")
            return d if d.is_finite() else _Dec("0")

        original = _bd(b.original_budget)
        revised = _bd(b.revised_budget)
        committed = _bd(b.committed)
        actual = _bd(b.actual)
        forecast = _bd(b.forecast_final)
        variance = revised - actual

        ws.cell(row=row_idx, column=3, value=original)
        ws.cell(row=row_idx, column=4, value=revised)
        ws.cell(row=row_idx, column=5, value=committed)
        ws.cell(row=row_idx, column=6, value=actual)
        ws.cell(row=row_idx, column=7, value=forecast)
        ws.cell(row=row_idx, column=8, value=variance)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="budgets_export.xlsx"'},
    )


@router.patch(
    "/budgets/{budget_id}",
    response_model=BudgetResponse,
    summary="Update budget line",
    description="Partially update a budget line. Only provided fields are modified.",
)
async def update_budget(
    budget_id: uuid.UUID,
    data: BudgetUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.update")),
    service: FinanceService = Depends(_get_service),
) -> BudgetResponse:
    """Update a budget line."""
    await _require_budget_access(session, budget_id, user_id)
    budget = await service.update_budget(budget_id, data)
    return BudgetResponse.model_validate(budget)


# ── EVM (MUST be before /{invoice_id}) ──────────────────────────────────────


@router.get(
    "/evm/",
    response_model=EVMListResponse,
    summary="List EVM snapshots",
    description="List Earned Value Management snapshots for a project. "
    "Each snapshot captures PV, EV, AC, SPI, CPI, and EAC at a point in time.",
)
async def list_evm_snapshots(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> EVMListResponse:
    """List EVM snapshots for a project."""
    await _require_project_access(session, project_id, user_id)
    # Scope an unfiltered listing to the caller's accessible projects so a
    # VIEWER does not receive every tenant's snapshots (admins -> None -> all).
    scope = None if project_id is not None else await accessible_project_ids(session, user_id)
    items, total = await service.list_evm_snapshots(project_id=project_id, project_ids=scope)
    return EVMListResponse(
        items=[EVMSnapshotResponse.model_validate(s) for s in items],
        total=total,
    )


@router.post(
    "/evm/snapshot/",
    response_model=EVMSnapshotResponse,
    status_code=201,
    summary="Create EVM snapshot",
    description="Capture a new Earned Value Management snapshot for a project. "
    "Records planned value (PV), earned value (EV), actual cost (AC), and derived indices.",
)
async def create_evm_snapshot(
    data: EVMSnapshotCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.create")),
    service: FinanceService = Depends(_get_service),
) -> EVMSnapshotResponse:
    """Create an EVM snapshot."""
    await _require_project_access(session, data.project_id, user_id)
    snapshot = await service.create_evm_snapshot(data)
    return EVMSnapshotResponse.model_validate(snapshot)


# ── Finance Dashboard ─────────────────────────────────────────────────────────


@router.get(
    "/dashboard/",
    summary="Get finance dashboard",
    description="Aggregated finance KPIs: payable, receivable, overdue totals, "
    "budget utilisation, cash flow overview, and budget warning level "
    "(normal / caution at 80%+ / critical at 95%+). "
    "Optionally scope to a single project via project_id query parameter.",
)
async def finance_dashboard(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> dict:
    """Aggregated finance KPIs: payable, receivable, overdue, budget, cash flow.

    Optionally scope to a single project via ``project_id`` query parameter.
    Returns budget warning level ("normal", "caution" at 80%+, "critical" at 95%+).
    """
    await _require_project_access(session, project_id, user_id)
    # Without an explicit project, aggregate only the caller's accessible
    # projects so a VIEWER's portfolio KPIs never sum across every tenant
    # (admins -> None -> full instance rollup).
    scope = None if project_id is not None else await accessible_project_ids(session, user_id)
    return await service.get_dashboard(project_id=project_id, project_ids=scope)


# ── ERP / accounting connectors (TOP-30 #4) ─────────────────────────────────
# Fixed-path block: declared BEFORE the parametric /{invoice_id} route so the
# /connectors/... segments are never parsed as an invoice UUID. Within the
# block, the literal /connectors/types/ and /connectors/logs/{log_id}/ routes
# come before /connectors/{config_id}/ for the same reason.


def _get_connector_service(session: SessionDep) -> ConnectorService:
    return ConnectorService(session)


@router.get(
    "/connectors/types/",
    response_model=list[ConnectorTypeInfo],
    summary="List available connector types",
    description="Catalogue of registered ERP / accounting connector types and their config fields.",
)
async def list_connector_types(
    _perm: None = Depends(RequirePermission("finance.connector.read")),
) -> list[ConnectorTypeInfo]:
    """Return the connector-type catalogue for the UI form builder."""
    return [ConnectorTypeInfo(**entry) for entry in connector_registry.list_types()]


@router.get(
    "/connectors/",
    response_model=ConnectorConfigListResponse,
    summary="List connectors",
    description="List configured connectors, optionally scoped to a project (also includes global connectors).",
)
async def list_connectors(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.connector.read")),
    service: ConnectorService = Depends(_get_connector_service),
) -> ConnectorConfigListResponse:
    """List connector configs for a project."""
    await _require_project_access(session, project_id, user_id)
    items = await service.list_configs(project_id=project_id)
    return ConnectorConfigListResponse(
        items=[ConnectorConfigResponse.from_model(c) for c in items],
        total=len(items),
    )


@router.post(
    "/connectors/",
    response_model=ConnectorConfigResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Create connector",
    description="Create a connector config. Credentials are encrypted at rest and never returned. MANAGER-only.",
)
async def create_connector(
    data: ConnectorConfigCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.connector.manage")),
    service: ConnectorService = Depends(_get_connector_service),
) -> ConnectorConfigResponse:
    """Create a connector configuration."""
    await _require_project_access(session, data.project_id, user_id)
    config = await service.create_config(data, actor_id=str(user_id) if user_id else None)
    return ConnectorConfigResponse.from_model(config)


@router.get(
    "/connectors/logs/{log_id}/",
    response_model=SyncLogResponse,
    summary="Get a sync-log entry",
    description="Retrieve a single connector sync-log entry by id.",
)
async def get_connector_log(
    log_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.connector.read")),
    service: ConnectorService = Depends(_get_connector_service),
) -> SyncLogResponse:
    """Get a single sync-log entry."""
    log = await service.get_log(log_id)
    if log is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sync log not found")
    await _require_project_access(session, log.project_id, user_id)
    return SyncLogResponse.from_model(log)


@router.get(
    "/connectors/{config_id}/",
    response_model=ConnectorConfigResponse,
    summary="Get connector",
    description="Retrieve a single connector config by id.",
)
async def get_connector(
    config_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.connector.read")),
    service: ConnectorService = Depends(_get_connector_service),
) -> ConnectorConfigResponse:
    """Get a single connector config."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    return ConnectorConfigResponse.from_model(config)


@router.patch(
    "/connectors/{config_id}/",
    response_model=ConnectorConfigResponse,
    summary="Update connector",
    description="Update a connector config. An absent or null credentials field leaves the stored secret unchanged. MANAGER-only.",
)
async def update_connector(
    config_id: uuid.UUID,
    data: ConnectorConfigUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.connector.manage")),
    service: ConnectorService = Depends(_get_connector_service),
) -> ConnectorConfigResponse:
    """Update a connector configuration."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    config = await service.update_config(config_id, data)
    return ConnectorConfigResponse.from_model(config)


@router.delete(
    "/connectors/{config_id}/",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete connector",
    description="Delete a connector config and its sync history. MANAGER-only.",
)
async def delete_connector(
    config_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.connector.manage")),
    service: ConnectorService = Depends(_get_connector_service),
) -> None:
    """Delete a connector configuration."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    await service.delete_config(config_id)


@router.post(
    "/connectors/{config_id}/validate/",
    response_model=ValidateResponse,
    summary="Validate connector config",
    description="Run the connector's config validation and return any problems. No side effects.",
)
async def validate_connector(
    config_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.connector.read")),
    service: ConnectorService = Depends(_get_connector_service),
) -> ValidateResponse:
    """Validate a connector config without running a sync."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    problems = await service.validate_config(config_id)
    return ValidateResponse(ok=not problems, problems=problems)


@router.post(
    "/connectors/{config_id}/sync/",
    response_model=SyncLogResponse,
    summary="Run a connector sync",
    description="Run a push and/or pull. dry_run=true previews with no side effects. "
    "A live run (dry_run=false) mutates external files and may write ledger rows. MANAGER-only.",
)
async def sync_connector(
    config_id: uuid.UUID,
    body: SyncTriggerRequest,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.connector.sync")),
    service: ConnectorService = Depends(_get_connector_service),
) -> SyncLogResponse:
    """Trigger a connector sync (manual)."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    # Rate-limit live syncs the same way as approve/pay; dry runs are cheap.
    if not body.dry_run:
        allowed, _ = approval_limiter.is_allowed(str(user_id))
        if not allowed:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail="Too many sync requests in a short time. Please wait a moment.",
            )
    log = await service.run_sync(
        config_id,
        direction=body.direction,
        dry_run=body.dry_run,
        actor_id=str(user_id) if user_id else None,
        trigger="manual",
    )
    return SyncLogResponse.from_model(log)


@router.get(
    "/connectors/{config_id}/logs/",
    response_model=SyncLogListResponse,
    summary="List sync history",
    description="List sync-log entries for a connector, newest first.",
)
async def list_connector_logs(
    config_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    _perm: None = Depends(RequirePermission("finance.connector.read")),
    service: ConnectorService = Depends(_get_connector_service),
) -> SyncLogListResponse:
    """List a connector's sync history."""
    config = await service.get_config_or_404(config_id)
    await _require_project_access(session, config.project_id, user_id)
    items, total = await service.list_logs(config_id=config_id, limit=limit, offset=offset)
    return SyncLogListResponse(items=[SyncLogResponse.from_model(x) for x in items], total=total)


# ── Invoice by ID (parametric routes LAST) ──────────────────────────────────


@router.get(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Get invoice",
    description="Retrieve a single invoice by its UUID, including line items and payment history.",
)
async def get_invoice(
    invoice_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("finance.read")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Get a single invoice by ID."""
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.get_invoice(invoice_id)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.patch(
    "/{invoice_id}",
    response_model=InvoiceResponse,
    summary="Update invoice",
    description="Partially update an invoice. Only provided fields are modified.",
)
async def update_invoice(
    invoice_id: uuid.UUID,
    data: InvoiceUpdate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.update")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Update an invoice."""
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.update_invoice(invoice_id, data)
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.post(
    "/{invoice_id}/approve/",
    response_model=InvoiceResponse,
    summary="Approve invoice",
    description="Transition an invoice to 'sent' (legacy alias 'approved') status. "
    "Only invoices in 'draft' or 'pending' status can be approved. "
    "MANAGER-only - invoice approval is the financial-commitment gate.",
)
async def approve_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.approve")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Approve an invoice.

    R7 (2026-05-24): pinned to ``finance.approve`` (MANAGER+).
    Previously this route used ``finance.update`` (EDITOR), which let
    any estimator commit a payable to the project.
    """
    allowed, _ = approval_limiter.is_allowed(str(user_id))
    if not allowed:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Rate limit exceeded. Try again later.")
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.approve_invoice(
        invoice_id,
        actor_id=str(user_id) if user_id else None,
    )
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


@router.post(
    "/{invoice_id}/pay/",
    response_model=InvoiceResponse,
    summary="Mark invoice as paid",
    description="Transition an invoice to 'paid' status. Records the payment date. "
    "MANAGER-only - marking an invoice paid is a binding ledger action. "
    "Idempotent: a second call against an already-paid invoice returns 400, "
    "not a duplicate ledger write.",
)
async def pay_invoice(
    invoice_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("finance.pay")),
    service: FinanceService = Depends(_get_service),
) -> InvoiceResponse:
    """Mark invoice as paid.

    R7 (2026-05-24): pinned to ``finance.pay`` (MANAGER+). The FSM
    allowlist (``_INVOICE_STATUS_TRANSITIONS``) ensures a second click
    against an already-paid invoice cannot re-trigger budget-actual
    recompute - it 400s on the disallowed ``paid -> paid`` transition
    (idempotency by allowlist).
    """
    allowed, _ = approval_limiter.is_allowed(str(user_id))
    if not allowed:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Rate limit exceeded. Try again later.")
    await _require_invoice_access(session, invoice_id, user_id)
    invoice = await service.pay_invoice(
        invoice_id,
        actor_id=str(user_id) if user_id else None,
    )
    names = await _fetch_counterparty_names(session, [invoice.contact_id])
    return _invoice_to_response(invoice, names)


# ── GAAP general ledger + financial reporting (task #77) ────────────────────
#
# Mounted under ``/gaap`` so the static segment never collides with the
# parametric ``/{invoice_id}`` route above. Endpoints:
#   GET    /gaap/accounts                  - list chart of accounts
#   POST   /gaap/accounts                  - create an account (MANAGER)
#   POST   /gaap/accounts/seed-defaults    - seed the default chart (MANAGER)
#   PATCH  /gaap/accounts/{id}             - update an account (MANAGER)
#   POST   /gaap/journal-entries           - post a balanced journal (MANAGER)
#   GET    /gaap/trial-balance             - trial balance (debits == credits)
#   GET    /gaap/statements/income         - income statement (P&L)
#   GET    /gaap/statements/balance-sheet  - balance sheet (A = L + E)
#   GET    /gaap/statements/cash-flow      - direct-method cash flow

gaap_router = APIRouter(prefix="/gaap", tags=["finance-gaap"])


def _ledger_entry_to_response(row: Any) -> LedgerEntryResponse:
    return LedgerEntryResponse.model_validate(row)


@gaap_router.get(
    "/accounts",
    response_model=LedgerAccountListResponse,
    summary="List chart of accounts",
    description="List chart-of-accounts rows for a scope. With a project_id the shared "
    "workspace chart is unioned with the project's own accounts.",
)
async def list_ledger_accounts(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    account_type: str | None = Query(default=None),
    active_only: bool = Query(default=False),
    _perm: None = Depends(RequirePermission("finance.gl.read")),
    service: FinanceService = Depends(_get_service),
) -> LedgerAccountListResponse:
    """List chart-of-accounts rows."""
    # project_id=None with include_workspace=True returns the company-wide
    # workspace chart unioned with every project's accounts, so this is the
    # consolidated GL view: admin-gate the None scope exactly like the GL
    # statements (a project_id falls through to the normal per-project check).
    await _require_gl_consolidated_scope(session, project_id, user_id)
    items, total = await service.list_accounts(
        project_id=project_id,
        account_type=account_type,
        active_only=active_only,
    )
    return LedgerAccountListResponse(
        items=[LedgerAccountResponse.model_validate(a) for a in items],
        total=total,
    )


@gaap_router.post(
    "/accounts",
    response_model=LedgerAccountResponse,
    status_code=201,
    summary="Create a ledger account",
    description="Create a chart-of-accounts account. normal_balance is derived from "
    "account_type when omitted. MANAGER-only - the chart shapes every statement.",
)
async def create_ledger_account(
    data: LedgerAccountCreate,
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.gl.manage_accounts")),
    service: FinanceService = Depends(_get_service),
) -> LedgerAccountResponse:
    """Create a chart-of-accounts account."""
    # A null project_id writes a workspace-level (company-wide) chart account,
    # so gate that path to admins like the consolidated GL; a project_id falls
    # through to the normal per-project owner check.
    await _require_gl_consolidated_scope(session, data.project_id, user_id)
    account = await service.create_account(data)
    return LedgerAccountResponse.model_validate(account)


@gaap_router.post(
    "/accounts/seed-defaults",
    response_model=LedgerAccountListResponse,
    summary="Seed the default construction chart of accounts",
    description="Idempotently seed the default chart into a scope (workspace when no project_id). MANAGER-only.",
)
async def seed_default_chart(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.gl.manage_accounts")),
    service: FinanceService = Depends(_get_service),
) -> LedgerAccountListResponse:
    """Seed the default chart of accounts."""
    # A null project_id seeds the workspace-level (company-wide) chart, so gate
    # that path to admins like the consolidated GL; a project_id falls through
    # to the normal per-project owner check.
    await _require_gl_consolidated_scope(session, project_id, user_id)
    accounts = await service.seed_default_chart(project_id=project_id)
    return LedgerAccountListResponse(
        items=[LedgerAccountResponse.model_validate(a) for a in accounts],
        total=len(accounts),
    )


@gaap_router.patch(
    "/accounts/{account_id}",
    response_model=LedgerAccountResponse,
    summary="Update a ledger account",
    description="Patch descriptive fields on a chart-of-accounts account. Code and type are immutable. MANAGER-only.",
)
async def update_ledger_account(
    account_id: uuid.UUID,
    data: LedgerAccountUpdate,
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.gl.manage_accounts")),
    service: FinanceService = Depends(_get_service),
) -> LedgerAccountResponse:
    """Update a chart-of-accounts account."""
    account = await service.get_account(account_id)
    # A workspace-level (company-wide) account has project_id=None, where the
    # per-project owner check no-ops. Gate that path to admins like the create
    # and seed endpoints (and the consolidated GL); a project-scoped account
    # falls through to the normal per-project owner check.
    await _require_gl_consolidated_scope(session, account.project_id, user_id)
    updated = await service.update_account(account_id, data)
    return LedgerAccountResponse.model_validate(updated)


@gaap_router.post(
    "/journal-entries",
    response_model=JournalEntryResponse,
    status_code=201,
    summary="Post a balanced journal entry",
    description="Post a balanced double-entry journal of two or more lines. The service "
    "rejects an unbalanced entry, a blended-currency entry, and posting to an "
    "unknown account. MANAGER-only - a binding ledger write.",
)
async def post_journal_entry(
    data: JournalEntryCreate,
    session: SessionDep,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("finance.gl.post_journal")),
    service: FinanceService = Depends(_get_service),
) -> JournalEntryResponse:
    """Post a balanced journal entry."""
    await _require_project_access(session, data.project_id, user_id)
    rows, total_debits, total_credits = await service.post_journal_entry(data)
    return JournalEntryResponse(
        transaction_ref=data.transaction_ref,
        lines=[_ledger_entry_to_response(r) for r in rows],
        total_debits=format(total_debits, "f"),
        total_credits=format(total_credits, "f"),
    )


@gaap_router.post(
    "/journal-entries/{transaction_ref}/reverse",
    response_model=LedgerTransactionResponse,
    status_code=201,
    summary="Reverse a posted journal transaction",
    description="Append a corrective reversal pair (accounts swapped, same magnitude) "
    "for a posted transaction. Idempotent: a transaction is reversible exactly "
    "once, so a repeat call returns the existing reversal rather than over-correcting "
    "the account. MANAGER-only - a binding ledger write.",
)
async def reverse_journal_entry(
    transaction_ref: str,
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project that owns the transaction being reversed."),
    description: str | None = Query(default=None, max_length=2000),
    _perm: None = Depends(RequirePermission("finance.gl.post_journal")),
    service: FinanceService = Depends(_get_service),
) -> LedgerTransactionResponse:
    """Reverse a posted ledger transaction by its reference."""
    await _require_project_access(session, project_id, user_id)
    rev_debit, rev_credit = await service.reverse_ledger_transaction(
        transaction_ref,
        project_id=project_id,
        description=description,
        created_by=user_id,
    )
    return LedgerTransactionResponse(
        debit=_ledger_entry_to_response(rev_debit),
        credit=_ledger_entry_to_response(rev_credit),
    )


@gaap_router.get(
    "/trial-balance",
    response_model=TrialBalanceResponse,
    summary="Trial balance",
    description="Per-account debit/credit totals and signed balance, with the grand "
    "debit == credit tie-out check, for a scope and period.",
)
async def get_trial_balance(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    currency_code: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.gl.read")),
    service: FinanceService = Depends(_get_service),
) -> TrialBalanceResponse:
    """Compute the trial balance."""
    await _require_gl_consolidated_scope(session, project_id, user_id)
    tb = await service.trial_balance(
        project_id=project_id,
        currency_code=currency_code,
        date_from=date_from,
        date_to=date_to,
    )
    return TrialBalanceResponse(
        currency=tb.currency,
        as_of=date_to,
        date_from=date_from,
        rows=[
            TrialBalanceRow(
                account_code=r.code,
                name=r.name,
                account_type=r.account_type.value,
                normal_balance=r.normal_balance.value,
                debit_total=format(r.debit_total, "f"),
                credit_total=format(r.credit_total, "f"),
                balance=format(r.signed_balance, "f"),
            )
            for r in tb.accounts
        ],
        total_debits=format(tb.total_debits, "f"),
        total_credits=format(tb.total_credits, "f"),
        is_balanced=tb.is_balanced,
        out_of_balance=format(tb.out_of_balance, "f"),
    )


def _statement_lines(lines: Any) -> list[StatementLineResponse]:
    return [
        StatementLineResponse(
            code=line.code,
            name=line.name,
            amount=format(line.amount, "f"),
            account_type=line.account_type,
            section=line.section,
        )
        for line in lines
    ]


@gaap_router.get(
    "/statements/income",
    response_model=IncomeStatementResponse,
    summary="Income statement (P&L)",
    description="Revenue minus expenses over a period, derived from the ledger.",
)
async def get_income_statement(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    currency_code: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.gl.read")),
    service: FinanceService = Depends(_get_service),
) -> IncomeStatementResponse:
    """Derive the income statement."""
    await _require_gl_consolidated_scope(session, project_id, user_id)
    inc = await service.income_statement(
        project_id=project_id,
        currency_code=currency_code,
        date_from=date_from,
        date_to=date_to,
    )
    return IncomeStatementResponse(
        currency=inc.currency,
        date_from=date_from,
        date_to=date_to,
        revenue_lines=_statement_lines(inc.revenue_lines),
        expense_lines=_statement_lines(inc.expense_lines),
        total_revenue=format(inc.total_revenue, "f"),
        total_expenses=format(inc.total_expenses, "f"),
        net_income=format(inc.net_income, "f"),
    )


@gaap_router.get(
    "/statements/balance-sheet",
    response_model=BalanceSheetResponse,
    summary="Balance sheet",
    description="Assets, liabilities and equity as of a date, with the assets = "
    "liabilities + equity tie-out check. Period net income is folded into "
    "equity as retained earnings so a live (pre-close) ledger still balances.",
)
async def get_balance_sheet(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    currency_code: str | None = Query(default=None),
    as_of: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.gl.read")),
    service: FinanceService = Depends(_get_service),
) -> BalanceSheetResponse:
    """Derive the balance sheet."""
    await _require_gl_consolidated_scope(session, project_id, user_id)
    bs = await service.balance_sheet(
        project_id=project_id,
        currency_code=currency_code,
        as_of=as_of,
    )
    return BalanceSheetResponse(
        currency=bs.currency,
        as_of=as_of,
        asset_lines=_statement_lines(bs.asset_lines),
        liability_lines=_statement_lines(bs.liability_lines),
        equity_lines=_statement_lines(bs.equity_lines),
        total_assets=format(bs.total_assets, "f"),
        total_liabilities=format(bs.total_liabilities, "f"),
        total_equity=format(bs.total_equity, "f"),
        liabilities_plus_equity=format(bs.liabilities_plus_equity, "f"),
        is_balanced=bs.is_balanced,
        out_of_balance=format(bs.out_of_balance, "f"),
    )


@gaap_router.get(
    "/statements/cash-flow",
    response_model=CashFlowResponse,
    summary="Cash flow statement (direct method)",
    description="Direct-method cash flow derived from movements on cash accounts, "
    "classified into operating / investing / financing by their counter-account. "
    "See the design doc for the honest scope of this derivation.",
)
async def get_cash_flow(
    session: SessionDep,
    user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    currency_code: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    _perm: None = Depends(RequirePermission("finance.gl.read")),
    service: FinanceService = Depends(_get_service),
) -> CashFlowResponse:
    """Derive the direct-method cash flow statement."""
    await _require_gl_consolidated_scope(session, project_id, user_id)
    cf = await service.cash_flow(
        project_id=project_id,
        currency_code=currency_code,
        date_from=date_from,
        date_to=date_to,
    )
    return CashFlowResponse(
        currency=cf.currency,
        date_from=date_from,
        date_to=date_to,
        method="direct",
        operating=format(cf.operating, "f"),
        investing=format(cf.investing, "f"),
        financing=format(cf.financing, "f"),
        opening_cash=format(cf.opening_cash, "f"),
        net_change=format(cf.net_change, "f"),
        closing_cash=format(cf.closing_cash, "f"),
        ties_out=cf.ties_out,
    )


# Mount the GAAP sub-router onto the module router so the loader picks it up.
router.include_router(gaap_router)


# Mount the invoice-approval DMS sub-router. Imported at the bottom on purpose:
# invoice_capture_router imports `_require_project_access` from this module, so
# the import must happen after that helper (and `router`) are defined to avoid a
# circular import at module load.
from app.modules.finance.invoice_capture_router import capture_router  # noqa: E402

router.include_router(capture_router)
