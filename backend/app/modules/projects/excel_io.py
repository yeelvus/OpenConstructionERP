# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Portfolio Excel import/export for projects.

Mirrors the contacts module pattern: one row = one project, flexible
header aliases (EN + 中文), openpyxl workbooks for template / export /
import. Coordinates may be decimal degrees or a combined DMS paste in
``lat`` (with ``lng`` empty).
"""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from app.modules.projects.schemas import ProjectCreate

# Canonical column order for template + export (stable round-trip).
PROJECT_EXCEL_COLUMNS: list[str] = [
    "name",
    "project_code",
    "description",
    "region",
    "classification_standard",
    "currency",
    "locale",
    "country_code",
    "project_type",
    "phase",
    "client_id",
    "contract_value",
    "budget_estimate",
    "contingency_pct",
    "gross_floor_area",
    "planned_start_date",
    "planned_end_date",
    "default_vat_rate",
    "street",
    "city",
    "state",
    "country",
    "postal_code",
    "lat",
    "lng",
]

# Header aliases → canonical key (matched case-insensitively after strip).
_COLUMN_ALIASES: dict[str, list[str]] = {
    "name": ["name", "project name", "project_name", "名称", "项目名称", "项目名"],
    "project_code": [
        "project_code",
        "code",
        "project code",
        "项目编号",
        "项目代码",
        "编号",
    ],
    "description": ["description", "desc", "描述", "说明", "项目描述"],
    "region": ["region", "市场", "区域", "地区"],
    "classification_standard": [
        "classification_standard",
        "classification",
        "standard",
        "分类标准",
        "清单标准",
    ],
    "currency": ["currency", "curr", "币种", "货币"],
    "locale": ["locale", "language", "lang", "语言", "界面语言"],
    "country_code": ["country_code", "country code", "iso country", "国家代码"],
    "project_type": ["project_type", "type", "类型", "项目类型"],
    "phase": ["phase", "阶段", "项目阶段"],
    "client_id": ["client_id", "client", "client name", "业主", "客户", "甲方"],
    "contract_value": ["contract_value", "contract value", "合同额", "合同金额"],
    "budget_estimate": ["budget_estimate", "budget", "预算", "预算金额"],
    "contingency_pct": [
        "contingency_pct",
        "contingency",
        "contingency %",
        "预备金比例",
        "不可预见费",
    ],
    "gross_floor_area": [
        "gross_floor_area",
        "gfa",
        "floor area",
        "建筑面积",
        "面积",
        "gfa_m2",
    ],
    "planned_start_date": [
        "planned_start_date",
        "start_date",
        "start",
        "计划开工",
        "开工日期",
        "开始日期",
    ],
    "planned_end_date": [
        "planned_end_date",
        "end_date",
        "end",
        "计划完工",
        "完工日期",
        "结束日期",
    ],
    "default_vat_rate": ["default_vat_rate", "vat", "vat_rate", "税率", "增值税"],
    "street": ["street", "address_street", "街道", "地址", "详细地址"],
    "city": ["city", "市", "城市"],
    "state": ["state", "province", "省", "州"],
    "country": ["country", "国家"],
    "postal_code": ["postal_code", "postcode", "zip", "邮编", "邮政编码"],
    "lat": ["lat", "latitude", "纬度", "gps_lat"],
    "lng": ["lng", "lon", "long", "longitude", "经度", "gps_lng"],
}

MAX_IMPORT_ROWS = 500

_XLSX_MAGIC = b"PK\x03\x04"
_CSV_BANNED_PREFIXES = (b"MZ", b"\x7fELF", b"\xca\xfe\xba\xbe", b"PK\x03\x04")

# Lightweight DMS / decimal pair parser (subset of frontend parseCoordinates).
_DMS_HEM = r"[NSnsEWew北南东西東]"
_DMS_PART = (
    rf"(?:{_DMS_HEM}\s*)?"
    rf"(-?\d+(?:\.\d+)?)\s*°?\s*"
    rf"(?:(\d+(?:\.\d+)?)\s*['′']?\s*)?"
    rf"(?:(\d+(?:\.\d+)?)\s*[\"″\"]?\s*)?"
    rf"(?:{_DMS_HEM})?"
)


def match_column(header: str) -> str | None:
    """Map a spreadsheet header to a canonical field name."""
    normalised = header.strip().lower()
    if not normalised:
        return None
    for canonical, aliases in _COLUMN_ALIASES.items():
        if normalised in {a.lower() for a in aliases}:
            return canonical
    # Exact canonical match
    if normalised in _COLUMN_ALIASES:
        return normalised
    return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        # Avoid 13.593780000000001 style noise from Excel numbers.
        if val == int(val):
            return str(int(val))
        return format(val, "f").rstrip("0").rstrip(".")
    if isinstance(val, Decimal):
        return format(val, "f").rstrip("0").rstrip(".")
    return str(val).strip()


def _empty_to_none(s: str) -> str | None:
    s = s.strip()
    return s if s else None


def _dms_to_decimal(deg: float, minutes: float, seconds: float, sign: float) -> float:
    return sign * (abs(deg) + minutes / 60.0 + seconds / 3600.0)


def _hem_sign(token: str, kind: str) -> float:
    t = token.upper()
    if kind == "lat":
        if t in ("S", "南"):
            return -1.0
        return 1.0
    if t in ("W", "西"):
        return -1.0
    return 1.0


def parse_coordinate_value(raw: str) -> float | None:
    """Parse a single coordinate cell (decimal or one DMS component)."""
    text = (raw or "").strip()
    if not text:
        return None
    # Pure decimal
    try:
        v = float(text.replace(",", "."))
        if abs(v) <= 180:
            return v
    except ValueError:
        pass

    # Strip hemisphere letters for decimal-with-hem: 13.5N
    m = re.match(rf"^({_DMS_HEM})?\s*(-?\d+(?:\.\d+)?)\s*({_DMS_HEM})?$", text)
    if m and not re.search(r"[°'\"]", text):
        try:
            v = float(m.group(2))
        except ValueError:
            return None
        hem = m.group(3) or m.group(1)
        if hem and hem.upper() in ("S", "南", "W", "西"):
            v = -abs(v)
        return v

    m = re.match(rf"^{_DMS_PART}$", text)
    if not m:
        return None
    try:
        deg = float(m.group(1))
        minutes = float(m.group(2) or 0)
        seconds = float(m.group(3) or 0)
    except (TypeError, ValueError):
        return None
    hems = re.findall(_DMS_HEM, text)
    sign = 1.0
    if hems:
        sign = _hem_sign(hems[-1], "lat" if hems[-1].upper() in "NS北南" else "lng")
    if deg < 0:
        sign = -1.0
        deg = abs(deg)
    return _dms_to_decimal(deg, minutes, seconds, sign)


def parse_lat_lng_pair(lat_raw: str, lng_raw: str) -> tuple[float | None, float | None]:
    """Resolve lat/lng cells; allow a full pair pasted into lat alone."""
    lat_s = (lat_raw or "").strip()
    lng_s = (lng_raw or "").strip()

    looks_like_pair = False
    if lat_s and not lng_s:
        hem_count = len(re.findall(_DMS_HEM, lat_s))
        deg_count = lat_s.count("°")
        if hem_count >= 2 or deg_count >= 2 or re.search(r"[,;]", lat_s):
            looks_like_pair = True
        elif re.match(r"^-?\d+(\.\d+)?\s+-?\d+(\.\d+)?$", lat_s):
            looks_like_pair = True

    if looks_like_pair:
        parts = re.split(r"\s*[,;]\s*", lat_s, maxsplit=1)
        if len(parts) == 2:
            return parse_coordinate_value(parts[0]), parse_coordinate_value(parts[1])
        hems = list(re.finditer(_DMS_HEM, lat_s))
        if len(hems) >= 2:
            mid = hems[0].end()
            return (
                parse_coordinate_value(lat_s[:mid]),
                parse_coordinate_value(lat_s[mid:]),
            )
        # Space-separated decimals or two ° groups without clean hems
        if re.match(r"^-?\d+(\.\d+)?\s+-?\d+(\.\d+)?$", lat_s):
            a, b = lat_s.split(None, 1)
            return parse_coordinate_value(a), parse_coordinate_value(b)
        # Split after first degree/minute/second cluster ending in hem or before second number°
        m = re.search(r"[°][^°]*?(?=\s+-?\d)", lat_s)
        if m:
            mid = m.end()
            # include trailing hem on first part if present just after m
            return (
                parse_coordinate_value(lat_s[:mid].strip()),
                parse_coordinate_value(lat_s[mid:].strip()),
            )

    lat = parse_coordinate_value(lat_s) if lat_s else None
    lng = parse_coordinate_value(lng_s) if lng_s else None
    return lat, lng


def row_to_project_create(row: dict[str, Any]) -> ProjectCreate:
    """Build a validated ``ProjectCreate`` from a normalised row dict.

    Raises ``ValueError`` with a human-readable message on validation failure.
    """
    name = _cell_str(row.get("name"))
    if not name:
        raise ValueError("name is required")

    lat, lng = parse_lat_lng_pair(_cell_str(row.get("lat")), _cell_str(row.get("lng")))
    if lat is not None and (lat < -90 or lat > 90):
        raise ValueError(f"lat out of range: {lat}")
    if lng is not None and (lng < -180 or lng > 180):
        raise ValueError(f"lng out of range: {lng}")
    if (lat is None) ^ (lng is None):
        raise ValueError("lat and lng must both be set or both empty")

    street = _empty_to_none(_cell_str(row.get("street")))
    city = _empty_to_none(_cell_str(row.get("city")))
    state = _empty_to_none(_cell_str(row.get("state")))
    country = _empty_to_none(_cell_str(row.get("country")))
    postal = _empty_to_none(_cell_str(row.get("postal_code")))

    address: dict[str, Any] | None = None
    if any([street, city, state, country, postal, lat is not None]):
        address = {
            "street": street,
            "city": city,
            "state": state,
            "country": country,
            "postal_code": postal,
            "lat": lat,
            "lng": lng,
        }

    payload: dict[str, Any] = {
        "name": name,
        "description": _cell_str(row.get("description")),
        "region": _cell_str(row.get("region")),
        "classification_standard": _cell_str(row.get("classification_standard")),
        "currency": _cell_str(row.get("currency")),
        "locale": _cell_str(row.get("locale")) or "en",
        "project_code": _empty_to_none(_cell_str(row.get("project_code"))),
        "project_type": _empty_to_none(_cell_str(row.get("project_type"))),
        "phase": _empty_to_none(_cell_str(row.get("phase"))),
        "client_id": _empty_to_none(_cell_str(row.get("client_id"))),
        "country_code": _empty_to_none(_cell_str(row.get("country_code"))),
        "contract_value": _empty_to_none(_cell_str(row.get("contract_value"))),
        "budget_estimate": _empty_to_none(_cell_str(row.get("budget_estimate"))),
        "contingency_pct": _empty_to_none(_cell_str(row.get("contingency_pct"))),
        "gross_floor_area": _empty_to_none(_cell_str(row.get("gross_floor_area"))),
        "planned_start_date": _empty_to_none(_cell_str(row.get("planned_start_date"))),
        "planned_end_date": _empty_to_none(_cell_str(row.get("planned_end_date"))),
        "default_vat_rate": _empty_to_none(_cell_str(row.get("default_vat_rate"))),
        "address": address,
    }

    try:
        return ProjectCreate.model_validate(payload)
    except Exception as exc:  # pydantic ValidationError
        # Flatten first error for the operator.
        msg = str(exc)
        if hasattr(exc, "errors"):
            try:
                errs = exc.errors()  # type: ignore[attr-defined]
                if errs:
                    e0 = errs[0]
                    loc = ".".join(str(x) for x in e0.get("loc", ()))
                    msg = f"{loc}: {e0.get('msg', msg)}" if loc else e0.get("msg", msg)
            except Exception:
                pass
        raise ValueError(msg) from exc


def project_to_export_row(project: Any) -> dict[str, Any]:
    """Flatten a Project ORM/response object to export columns."""
    addr = getattr(project, "address", None) or {}
    if not isinstance(addr, dict):
        addr = {}

    def g(key: str, default: str = "") -> str:
        val = getattr(project, key, None)
        if val is None:
            return default
        return str(val)

    return {
        "name": g("name"),
        "project_code": g("project_code"),
        "description": g("description"),
        "region": g("region"),
        "classification_standard": g("classification_standard"),
        "currency": g("currency"),
        "locale": g("locale", "en"),
        "country_code": g("country_code"),
        "project_type": g("project_type"),
        "phase": g("phase"),
        "client_id": g("client_id"),
        "contract_value": g("contract_value"),
        "budget_estimate": g("budget_estimate"),
        "contingency_pct": g("contingency_pct"),
        "gross_floor_area": g("gross_floor_area"),
        "planned_start_date": g("planned_start_date"),
        "planned_end_date": g("planned_end_date"),
        "default_vat_rate": g("default_vat_rate"),
        "street": addr.get("street") or "",
        "city": addr.get("city") or "",
        "state": addr.get("state") or "",
        "country": addr.get("country") or "",
        "postal_code": addr.get("postal_code") or "",
        "lat": addr.get("lat") if addr.get("lat") is not None else "",
        "lng": addr.get("lng") if addr.get("lng") is not None else "",
    }


def sniff_upload(filename: str, content: bytes) -> str:
    """Return ``xlsx`` or ``csv``; raise ValueError on mismatch."""
    name = (filename or "").lower()
    head = content[:8]
    if name.endswith(".xlsx"):
        if not head.startswith(_XLSX_MAGIC):
            raise ValueError("File does not look like a valid .xlsx (missing ZIP signature).")
        return "xlsx"
    if name.endswith(".csv"):
        for sig in _CSV_BANNED_PREFIXES:
            if head.startswith(sig):
                raise ValueError("File does not look like CSV (binary signature detected).")
        return "csv"
    raise ValueError("Unsupported file type. Upload .xlsx or .csv.")


def parse_rows_from_excel(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no worksheets")
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            raise ValueError("Excel file is empty or has no header row")
        column_map: dict[int, str] = {}
        for idx, hdr in enumerate(raw_headers):
            if hdr is not None:
                canonical = match_column(str(hdr))
                if canonical:
                    column_map[idx] = canonical
        if "name" not in column_map.values():
            raise ValueError(
                "Missing required column 'name' (or 名称 / 项目名称). "
                "Download the template for the correct headers."
            )
        rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            row: dict[str, Any] = {}
            for idx, val in enumerate(raw_row):
                canonical = column_map.get(idx)
                if canonical and val is not None and str(val).strip() != "":
                    row[canonical] = val
            if row:
                rows.append(row)
        return rows
    finally:
        wb.close()


def parse_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file — unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = match_column(hdr)
        if canonical:
            column_map[idx] = canonical
    if "name" not in column_map.values():
        raise ValueError("Missing required column 'name' (or 名称 / 项目名称).")

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None and str(val).strip() != "":
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)
    return rows


def build_template_workbook() -> bytes:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Projects"

    headers = PROJECT_EXCEL_COLUMNS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    comments = {
        "name": "Required. Project display name.",
        "project_code": "Optional unique code. Leave blank to auto-generate PRJ-YYYY-NNNN.",
        "lat": "Decimal degrees or DMS (e.g. 13.59378 or 13°35'37.60\"N).",
        "lng": "Decimal degrees or DMS (e.g. 100.96344 or 100°57'48.38\"E).",
        "country_code": "ISO 3166-1 alpha-2 (e.g. TH, CN, DE).",
        "currency": "ISO 4217 (e.g. THB, CNY, EUR). Leave empty if undecided.",
        "planned_start_date": "Prefer YYYY-MM-DD.",
    }
    for col_idx, key in enumerate(headers, start=1):
        if key in comments:
            ws.cell(row=1, column=col_idx).comment = Comment(comments[key], "OpenConstructionERP")

    # Example rows (not imported if user deletes them — they are real data if left)
    ws.append(
        [
            "Example Site Alpha",
            "THCC-DEMO-001",
            "Sample row — replace or delete before import",
            "SoutheastAsia",
            "",
            "THB",
            "en",
            "TH",
            "Industrial",
            "",
            "",
            "",
            "",
            "",
            "",
            "2026-01-01",
            "2027-12-31",
            "",
            "",
            "Chonburi",
            "",
            "Thailand",
            "",
            13.59378,
            100.96344,
        ]
    )
    ws.append(
        [
            "示例项目 Beta",
            "",
            "可删除的示例行；名称必填",
            "China",
            "",
            "CNY",
            "zh",
            "CN",
            "住宅",
            "",
            "",
            "",
            "",
            "",
            "12000",
            "",
            "",
            "",
            "",
            "邯郸",
            "河北",
            "中国",
            "",
            "",
            "",
        ]
    )

    # Second sheet: column legend in Chinese + English
    legend = wb.create_sheet("说明_Columns")
    legend.append(["column", "中文", "required", "notes"])
    legend["A1"].font = Font(bold=True)
    legend_rows = [
        ("name", "名称", "yes", "Project name"),
        ("project_code", "项目编号", "no", "Unique; auto if empty"),
        ("description", "描述", "no", ""),
        ("region", "区域/市场", "no", "China, SoutheastAsia, DACH, …"),
        ("currency", "币种", "no", "CNY, THB, EUR, …"),
        ("country_code", "国家代码", "no", "ISO-2"),
        ("street/city/…", "地址字段", "no", "→ address JSON"),
        ("lat/lng", "纬度/经度", "no", "Decimal or DMS; both or neither"),
    ]
    for r in legend_rows:
        legend.append(list(r))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_workbook(projects: list[Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Projects"
    ws.append(PROJECT_EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for project in projects:
        row = project_to_export_row(project)
        ws.append([row.get(col, "") for col in PROJECT_EXCEL_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
