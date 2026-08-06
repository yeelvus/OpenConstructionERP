# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
"""THCC local payment-ledger import (B_财务付款_std).

Scans a folder of per-project Excel workbooks::

    …/12-财务数据💰/A_财务付款数据/B_财务付款_std/
        THCC-2026-004_天顿科技_(2026-07-31).xlsx
        …

Each workbook's ``明细数据`` sheet (preferred over ``明细数据_万泰铢``)
contains payment rows. Import creates one payable invoice + one payment
per non-zero 实付金额 row, with a stable idempotency key so re-import
is safe.

No file is copied into application storage — only amounts / metadata
are written to the finance tables.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import uuid
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

# Host Mac default (when running outside Docker).
_HOST_DEFAULT_ROOT = (
    Path.home()
    / "Desktop"
    / "邯郸中材"
    / "01成本统计"
    / "12-财务数据💰"
    / "A_财务付款数据"
    / "B_财务付款_std"
)

# Docker bind-mount default (see docker-compose.custom.yml → /host/thcc).
_DOCKER_DEFAULT_ROOT = Path(
    "/host/thcc/01成本统计/12-财务数据💰/A_财务付款数据/B_财务付款_std"
)


def _config_file() -> Path:
    """Prefer OE_DATA_DIR so Docker volume keeps the path setting."""
    data = (os.environ.get("OE_DATA_DIR") or "").strip()
    if data:
        return Path(data) / "thcc_payments.json"
    return Path.home() / ".openestimate" / "thcc_payments.json"

# THCC-2026-004 or embedded in filename
_PROJECT_CODE_RE = re.compile(r"(THCC-\d{4}-\d{3}|LSBM-\d{2}|XCY-\d{2})", re.I)

# Preferred sheet (local currency units, not 万)
_SHEET_PREFERRED = ("明细数据", "明细数据_万泰铢", "Sheet1")

# Header aliases (first matching column wins)
_HEADERS = {
    "project_code": ("项目编码", "project_code", "编码"),
    "project_name": ("项目名称", "project_name", "名称"),
    "account": ("科目",),
    "l1": ("一级科目",),
    "l2": ("二级科目",),
    "l3": ("三级科目",),
    "category_id": ("分类ID", "分类id"),
    "summary": ("摘要", "说明", "备注"),
    "supplier": ("供应商名称", "供应商", "对方单位", "收款单位"),
    "payable": ("应付金额", "应付"),
    "withholding": ("预扣税额", "预扣税", "代扣税"),
    "paid": ("实付金额", "实付", "付款金额", "金额"),
    "vat": ("增值税额", "增值税"),
    "net": ("不含税额", "不含税"),
    "pay_date": ("日期", "付款日期", "支付日期"),
    "month": ("月",),
    "year": ("年",),
}


def get_payments_root() -> Path:
    """Return configured root, falling back to Docker mount then host Desktop.

    Priority:
      1. ``THCC_PAYMENTS_ROOT`` env (set by docker-compose.custom.yml)
      2. Persisted config under OE_DATA_DIR / ~/.openestimate
      3. Docker bind-mount ``/host/thcc/...`` if present
      4. Host ``~/Desktop/邯郸中材/...``
    """
    env = (os.environ.get("THCC_PAYMENTS_ROOT") or "").strip()
    if env:
        return Path(env).expanduser()
    try:
        cfg = _config_file()
        if cfg.is_file():
            import json

            data = json.loads(cfg.read_text(encoding="utf-8"))
            raw = (data.get("root") or "").strip()
            if raw:
                return Path(raw).expanduser()
    except Exception:  # noqa: BLE001
        logger.debug("Could not read thcc_payments config", exc_info=True)
    if _DOCKER_DEFAULT_ROOT.is_dir():
        return _DOCKER_DEFAULT_ROOT
    return _HOST_DEFAULT_ROOT


def set_payments_root(root: str | Path) -> Path:
    path = Path(root).expanduser().resolve()
    cfg = _config_file()
    cfg.parent.mkdir(parents=True, exist_ok=True)
    import json

    cfg.write_text(
        json.dumps({"root": str(path), "updated_at": datetime.now(UTC).isoformat()}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _cell_decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value if value.is_finite() else Decimal("0")
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return Decimal("0")
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text or text in {"-", "—", "–"}:
        return Decimal("0")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _normalize_date(value: object, year: object = None, month: object = None) -> str | None:
    """Return YYYY-MM-DD or None."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _cell_str(value)
    if text:
        # 2026-04-27 / 2026/4/27 / 2026.04.27
        m = re.match(r"^(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})", text)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return date(y, mo, d).isoformat()
            except ValueError:
                pass
        # Excel serial
        try:
            serial = float(text)
            if 20000 < serial < 80000:
                from datetime import timedelta

                base = date(1899, 12, 30)
                return (base + timedelta(days=int(serial))).isoformat()
        except (TypeError, ValueError):
            pass
    # fallback year+month → first of month
    try:
        y = int(float(str(year))) if year not in (None, "") else None
        mo = int(float(str(month))) if month not in (None, "") else None
        if y and mo:
            return date(y, mo, 1).isoformat()
    except (TypeError, ValueError):
        pass
    return None


def _header_index_map(header_row: tuple[object, ...] | list[object]) -> dict[str, int]:
    labels = {_cell_str(h): i for i, h in enumerate(header_row) if _cell_str(h)}
    out: dict[str, int] = {}
    for key, aliases in _HEADERS.items():
        for alias in aliases:
            if alias in labels:
                out[key] = labels[alias]
                break
    return out


def _row_get(row: tuple[object, ...], idx_map: dict[str, int], key: str) -> object:
    i = idx_map.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


def project_code_from_filename(name: str) -> str | None:
    m = _PROJECT_CODE_RE.search(name)
    return m.group(1).upper() if m else None


def project_name_hint_from_filename(name: str) -> str:
    """THCC-2026-004_天顿科技_(2026-07-31).xlsx → 天顿科技"""
    stem = Path(name).stem
    # strip trailing _(date)
    stem = re.sub(r"_\(\d{4}-\d{2}-\d{2}\)$", "", stem)
    m = re.match(r"^(?:THCC-\d{4}-\d{3}|LSBM-\d{2}|XCY-\d{2}|未识别编码|需手动核对)_(.+)$", stem, re.I)
    if m:
        return m.group(1).strip()
    if "_" in stem:
        return stem.split("_", 1)[1].strip()
    return stem


@dataclass
class PaymentRow:
    project_code: str
    project_name: str
    account: str
    l1: str
    l2: str
    l3: str
    category_id: str
    summary: str
    supplier: str
    payable: Decimal
    withholding: Decimal
    paid: Decimal
    vat: Decimal
    net: Decimal
    pay_date: str
    source_file: str
    row_number: int

    def idempotency_key(self) -> str:
        raw = "|".join(
            [
                self.project_code,
                self.pay_date,
                self.supplier,
                self.summary,
                self.account,
                self.l2,
                self.category_id,
                format(self.paid, "f"),
                format(self.payable, "f"),
                str(self.row_number),
                self.source_file,
            ]
        )
        digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]
        return f"thccpay-{digest}"

    def invoice_number(self) -> str:
        # max 50 chars on invoice_number column
        digest = self.idempotency_key().replace("thccpay-", "")[:12]
        return f"THCC-PAY-{digest}"


@dataclass
class DiscoveredPaymentFile:
    filename: str
    path: str
    project_code: str | None
    project_name_hint: str
    row_count: int = 0
    paid_row_count: int = 0
    total_paid: str = "0"
    project_id: str | None = None
    project_matched: bool = False
    error: str | None = None


def parse_payment_workbook(path: Path) -> list[PaymentRow]:
    """Parse one std workbook into PaymentRow list (non-zero 实付 only)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import payment Excel files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = None
        for candidate in _SHEET_PREFERRED:
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if not sheet_name:
            return []
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        idx = _header_index_map(header)
        if "paid" not in idx and "payable" not in idx:
            # wrong sheet
            return []

        file_code = project_code_from_filename(path.name) or ""
        out: list[PaymentRow] = []
        for n, row in enumerate(rows_iter, start=2):
            if not row or not any(row):
                continue
            paid = _cell_decimal(_row_get(row, idx, "paid"))
            payable = _cell_decimal(_row_get(row, idx, "payable"))
            if paid <= 0 and payable <= 0:
                continue
            if paid <= 0:
                paid = payable
            pay_date = _normalize_date(
                _row_get(row, idx, "pay_date"),
                year=_row_get(row, idx, "year"),
                month=_row_get(row, idx, "month"),
            )
            if not pay_date:
                pay_date = date.today().isoformat()
            code = _cell_str(_row_get(row, idx, "project_code")).upper() or file_code
            out.append(
                PaymentRow(
                    project_code=code,
                    project_name=_cell_str(_row_get(row, idx, "project_name")),
                    account=_cell_str(_row_get(row, idx, "account")),
                    l1=_cell_str(_row_get(row, idx, "l1")),
                    l2=_cell_str(_row_get(row, idx, "l2")),
                    l3=_cell_str(_row_get(row, idx, "l3")),
                    category_id=_cell_str(_row_get(row, idx, "category_id")),
                    summary=_cell_str(_row_get(row, idx, "summary")),
                    supplier=_cell_str(_row_get(row, idx, "supplier")) or "未注明供应商",
                    payable=payable if payable > 0 else paid,
                    withholding=_cell_decimal(_row_get(row, idx, "withholding")),
                    paid=paid,
                    vat=_cell_decimal(_row_get(row, idx, "vat")),
                    net=_cell_decimal(_row_get(row, idx, "net")),
                    pay_date=pay_date,
                    source_file=path.name,
                    row_number=n,
                )
            )
        return out
    finally:
        wb.close()


def scan_payment_files(root: Path | None = None) -> list[DiscoveredPaymentFile]:
    """List xlsx files under the payments root with light stats."""
    root = root or get_payments_root()
    if not root.is_dir():
        return []
    discovered: list[DiscoveredPaymentFile] = []
    for path in sorted(root.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        code = project_code_from_filename(path.name)
        hint = project_name_hint_from_filename(path.name)
        entry = DiscoveredPaymentFile(
            filename=path.name,
            path=str(path),
            project_code=code,
            project_name_hint=hint,
        )
        try:
            rows = parse_payment_workbook(path)
            entry.row_count = len(rows)
            entry.paid_row_count = sum(1 for r in rows if r.paid > 0)
            total = sum((r.paid for r in rows), Decimal("0"))
            entry.total_paid = format(total, "f")
        except Exception as exc:  # noqa: BLE001
            entry.error = str(exc)[:200]
            logger.warning("Failed to parse payment file %s: %s", path.name, exc)
        discovered.append(entry)
    return discovered


async def _load_project_maps(session: AsyncSession) -> tuple[dict[str, Any], dict[str, Any]]:
    """Return (by_code_upper, by_name_norm) project maps (non-archived preferred)."""
    from app.modules.projects.models import Project

    rows = (await session.execute(select(Project))).scalars().all()
    by_code: dict[str, Any] = {}
    by_name: dict[str, Any] = {}
    for p in rows:
        code = (p.project_code or "").strip().upper()
        if code:
            existing = by_code.get(code)
            if existing is None or (
                (existing.status or "") == "archived" and (p.status or "") != "archived"
            ):
                by_code[code] = p
        name_key = " ".join((p.name or "").casefold().split())
        if name_key:
            existing = by_name.get(name_key)
            if existing is None or (
                (existing.status or "") == "archived" and (p.status or "") != "archived"
            ):
                by_name[name_key] = p
    return by_code, by_name


async def enrich_scan_with_projects(
    session: AsyncSession,
    files: list[DiscoveredPaymentFile],
) -> list[dict[str, Any]]:
    by_code, by_name = await _load_project_maps(session)
    out: list[dict[str, Any]] = []
    for f in files:
        d = asdict(f)
        proj = None
        if f.project_code:
            proj = by_code.get(f.project_code.upper())
        if proj is None and f.project_name_hint:
            key = " ".join(f.project_name_hint.casefold().split())
            proj = by_name.get(key)
        if proj is not None:
            d["project_id"] = str(proj.id)
            d["project_matched"] = True
            d["project_name"] = proj.name
            d["currency"] = proj.currency or ""
        else:
            d["project_matched"] = False
        out.append(d)
    return out


async def import_payment_rows(
    session: AsyncSession,
    *,
    rows: list[PaymentRow],
    project_id: uuid.UUID,
    currency_code: str,
    dry_run: bool = False,
    actor_id: str | None = None,
) -> dict[str, Any]:
    """Import payment rows into finance for one project.

    Creates payable invoice (status=paid) + payment per row. Skips when
    payment idempotency_key already exists.
    """
    from app.modules.finance.models import Invoice, InvoiceLineItem, Payment
    from app.modules.finance.repository import PaymentRepository

    pay_repo = PaymentRepository(session)
    imported = 0
    skipped = 0
    skipped_duplicate = 0
    errors: list[dict[str, Any]] = []
    total_amount = Decimal("0")

    for row in rows:
        if row.paid <= 0:
            skipped += 1
            continue
        key = row.idempotency_key()
        existing = await pay_repo.get_by_idempotency_key(key)
        if existing is not None:
            skipped_duplicate += 1
            skipped += 1
            continue

        inv_no = row.invoice_number()
        # Also skip if invoice number already taken for this project
        from app.modules.finance.repository import InvoiceRepository

        inv_repo = InvoiceRepository(session)
        if await inv_repo.invoice_number_taken(project_id, "payable", inv_no):
            skipped_duplicate += 1
            skipped += 1
            continue

        if dry_run:
            imported += 1
            total_amount += row.paid
            continue

        try:
            tax = row.vat if row.vat > 0 else Decimal("0")
            subtotal = row.net if row.net > 0 else (row.payable - tax if row.payable > tax else row.paid)
            if subtotal <= 0:
                subtotal = row.paid
            desc = row.summary or row.account or row.l2 or "付款"
            if row.supplier:
                notes = f"{row.supplier} · {desc}"
            else:
                notes = desc
            category = row.l2 or row.account or row.l1 or None
            if category and len(category) > 100:
                category = category[:100]

            meta = {
                "source": "thcc_payment_std",
                "source_file": row.source_file,
                "source_row": row.row_number,
                "project_code": row.project_code,
                "account": row.account,
                "l1": row.l1,
                "l2": row.l2,
                "l3": row.l3,
                "category_id": row.category_id,
                "supplier": row.supplier,
                "summary": row.summary,
                "payable": format(row.payable, "f"),
                "withholding": format(row.withholding, "f"),
                "vat": format(row.vat, "f"),
            }

            invoice = Invoice(
                project_id=project_id,
                contact_id=None,
                invoice_direction="payable",
                invoice_number=inv_no,
                invoice_date=row.pay_date,
                due_date=row.pay_date,
                currency_code=currency_code or "",
                amount_subtotal=subtotal,
                tax_amount=tax,
                retention_amount=Decimal("0"),
                amount_total=subtotal + tax if (subtotal + tax) > 0 else row.paid,
                status="paid",
                notes=notes[:5000],
                created_by=uuid.UUID(actor_id) if actor_id else None,
                metadata_=meta,
            )
            session.add(invoice)
            await session.flush()

            line = InvoiceLineItem(
                invoice_id=invoice.id,
                description=(desc[:500] if desc else "付款"),
                quantity=Decimal("1"),
                unit=None,
                unit_rate=row.paid,
                amount=row.paid,
                cost_category=category,
                sort_order=0,
            )
            session.add(line)

            withholding = row.withholding if row.withholding > 0 else Decimal("0")
            payment = Payment(
                invoice_id=invoice.id,
                payment_date=row.pay_date,
                amount=row.paid,
                currency_code=currency_code or "",
                exchange_rate_snapshot=Decimal("1"),
                reference=f"{row.supplier[:80]} · {desc[:80]}"[:255],
                idempotency_key=key,
                is_refund=False,
                withholding_amount=withholding,
                metadata_=meta,
            )
            session.add(payment)
            await session.flush()
            imported += 1
            total_amount += row.paid
        except Exception as exc:  # noqa: BLE001
            logger.exception("Failed to import payment row %s from %s", row.row_number, row.source_file)
            errors.append(
                {
                    "row": row.row_number,
                    "file": row.source_file,
                    "error": str(exc)[:200],
                    "summary": row.summary,
                }
            )

    # Caller session commits at end of request (do not commit here).

    return {
        "imported": imported,
        "skipped": skipped,
        "skipped_duplicate": skipped_duplicate,
        "errors": errors[:50],
        "error_count": len(errors),
        "total_paid": format(total_amount, "f"),
        "dry_run": dry_run,
    }


async def import_from_local_root(
    session: AsyncSession,
    *,
    project_id: uuid.UUID | None = None,
    project_codes: list[str] | None = None,
    filenames: list[str] | None = None,
    dry_run: bool = False,
    actor_id: str | None = None,
    root: Path | None = None,
) -> dict[str, Any]:
    """Scan local root and import matching workbooks."""
    root = root or get_payments_root()
    if not root.is_dir():
        return {
            "ok": False,
            "error": f"Payments root not found: {root}",
            "root": str(root),
            "files": 0,
            "imported": 0,
        }

    by_code, by_name = await _load_project_maps(session)
    files = scan_payment_files(root)
    code_filter = {c.strip().upper() for c in (project_codes or []) if c and c.strip()}
    name_filter = {f.strip() for f in (filenames or []) if f and f.strip()}

    results: list[dict[str, Any]] = []
    grand = {
        "imported": 0,
        "skipped": 0,
        "skipped_duplicate": 0,
        "error_count": 0,
        "files_processed": 0,
        "files_skipped": 0,
    }

    for f in files:
        if name_filter and f.filename not in name_filter:
            continue
        if code_filter and (not f.project_code or f.project_code.upper() not in code_filter):
            continue

        proj = None
        if project_id is not None:
            # Only import rows for this project; match file by project's code
            from app.modules.projects.models import Project

            proj = await session.get(Project, project_id)
            if proj is None:
                continue
            pcode = (proj.project_code or "").strip().upper()
            if f.project_code and pcode and f.project_code.upper() != pcode:
                # also allow name match for 未识别编码 files
                hint = " ".join(f.project_name_hint.casefold().split())
                pname = " ".join((proj.name or "").casefold().split())
                if hint != pname:
                    continue
        else:
            if f.project_code:
                proj = by_code.get(f.project_code.upper())
            if proj is None and f.project_name_hint:
                proj = by_name.get(" ".join(f.project_name_hint.casefold().split()))

        if proj is None:
            grand["files_skipped"] += 1
            results.append(
                {
                    "filename": f.filename,
                    "project_code": f.project_code,
                    "status": "unmatched",
                    "error": "No matching project in OCE",
                }
            )
            continue

        try:
            rows = parse_payment_workbook(Path(f.path))
        except Exception as exc:  # noqa: BLE001
            grand["files_skipped"] += 1
            results.append(
                {
                    "filename": f.filename,
                    "status": "parse_error",
                    "error": str(exc)[:200],
                }
            )
            continue

        # When scoped to a project, only keep rows for that code (or all if file matched)
        if project_id is not None and proj.project_code:
            pcode = proj.project_code.strip().upper()
            filtered = [r for r in rows if not r.project_code or r.project_code.upper() == pcode]
            if filtered:
                rows = filtered

        currency = (proj.currency or "").strip() or "THB"
        stats = await import_payment_rows(
            session,
            rows=rows,
            project_id=proj.id,
            currency_code=currency,
            dry_run=dry_run,
            actor_id=actor_id,
        )
        grand["imported"] += stats["imported"]
        grand["skipped"] += stats["skipped"]
        grand["skipped_duplicate"] += stats["skipped_duplicate"]
        grand["error_count"] += stats["error_count"]
        grand["files_processed"] += 1
        results.append(
            {
                "filename": f.filename,
                "project_id": str(proj.id),
                "project_code": proj.project_code,
                "project_name": proj.name,
                "status": "ok",
                **{k: stats[k] for k in ("imported", "skipped", "skipped_duplicate", "total_paid", "error_count")},
            }
        )

    return {
        "ok": True,
        "root": str(root),
        "dry_run": dry_run,
        "summary": grand,
        "files": results,
    }
